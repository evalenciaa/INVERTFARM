"""
farmacia/views/pdf_views.py
Vistas para generar reportes en PDF y Excel desde entradas (registro manual/carga masiva).
"""
import json
import logging
import os
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

logger = logging.getLogger(__name__)


@csrf_exempt
@login_required
def generar_reporte_pdf(request):    
    try:
        data = json.loads(request.body)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                              rightMargin=20*mm, leftMargin=20*mm,
                              topMargin=15*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        elements = []
        
        medicamento_style = ParagraphStyle(
            name='MedicamentoStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=9,
            alignment=0,
            wordWrap='LTR',
            splitLongWords=True
        )
        
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia/static/farmacia/img/logo.jpg')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=180*mm, height=(175*180/1236)*mm)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 15*mm))
        
        titulo = Paragraph('<para align=center><font size=14><b>REPORTE DE ENTRADA DE MEDICAMENTOS</b></font></para>', styles['Normal'])
        elements.append(titulo)
        elements.append(Spacer(1, 8*mm))
        
        info_data = [
            ['Folio: ', data.get('folio', 'N/A'), 'Fecha: ', data.get('fecha', 'N/A')],
            ['Tipo Entrada: ', data.get('tipo_entrada', 'N/A'), 'Almacén: ', data.get('almacen_nombre', 'N/A')],
            ['Fuente Financ.: ', data.get('fuente_financiamiento_nombre', 'N/A'), 'Proceso: ', data.get('proceso', 'N/A')]
        ]
        
        info_table = Table(info_data, colWidths=[35*mm, 60*mm, 40*mm, 60*mm])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 8*mm))
        
        encabezados = [
            Paragraph('<b>Medicamento</b>', styles['Normal']),
            Paragraph('<b>Lote</b>', styles['Normal']),
            Paragraph('<b>Presentación</b>', styles['Normal']),
            Paragraph('<b>Cantidad</b>', styles['Normal']),
            Paragraph('<b>P. Unitario</b>', styles['Normal']),
            Paragraph('<b>Total</b>', styles['Normal'])
        ]
        datos_tabla = [encabezados]

        for item in data.get('items', []):
            fila = [
                Paragraph(item.get('nombre', '')), 
                item.get('lote', ''),
                item.get('presentacion', ''),
                str(item.get('cantidad', 0)),
                f"${float(item.get('precio_unitario', 0)):,.2f}",
                f"${float(item.get('total', 0)):,.2f}"
            ]
            datos_tabla.append(fila)

        datos_tabla.append([
            '', '', '', '',
            Paragraph('<b>TOTAL GENERAL:</b>', styles['Normal']),
            f"${float(data.get('total', 0)):,.2f}"
        ])
        
        tabla = Table(datos_tabla, colWidths=[80*mm, 25*mm, 35*mm, 20*mm, 25*mm, 25*mm])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (4,-1), (-1,-1), colors.HexColor('#70AD47')),
            ('TEXTCOLOR', (4,-1), (-1,-1), colors.whitesmoke),
            ('FONTNAME', (4,-1), (-1,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('FONTSIZE', (0,1), (-1,-2), 8),
            ('LEADING', (0,1), (0,-2), 9),
        ]))
        
        elements.append(tabla)
        elements.append(Spacer(1, 15*mm))
        
        firmas_data = [
            ['', '', ''],
            ['________________________', '________________________', '________________________'],
            ['Recibido por', 'Autorizado por', 'Entregado por'],
            ['Nombre y Firma', 'Nombre y Firma', 'Nombre y Firma']
        ]
        
        firmas_table = Table(firmas_data, colWidths=[60*mm, 60*mm, 60*mm])
        firmas_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('SPAN', (0,0), (2,0)),
            ('LEADING', (1,1), (2,1), 14),
        ]))
        elements.append(firmas_table)
        
        elements.append(Spacer(1, 10*mm))
        footer = Paragraph(
            f"<font size=7>Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} por {request.user.get_full_name()} | Sistema de Gestión Farmacéutica</font>", 
            styles['Normal'])
        elements.append(footer)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="ENTRADA_{data.get("folio", "REPORTE")}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Error al generar PDF: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)
                

@csrf_exempt
@login_required
def generar_reporte_excel(request):
    try:
        data = json.loads(request.body)
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Entradas"

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        logo_path = os.path.join(settings.BASE_DIR, 'farmacia', 'static', 'farmacia', 'img', 'logo.jpg')
        if os.path.exists(logo_path):
            img = ExcelImage(logo_path)
            img.width = 1236 * 0.75
            img.height = 175 * 0.75
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 135

        ws.merge_cells('A3:F3')
        title_cell = ws['A3']
        title_cell.value = "REPORTE DE ENTRADA DE MEDICAMENTOS"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        header_data = [
            ("Folio:", data.get('folio', '')),
            ("Fecha:", data.get('fecha', '')),
            ("Tipo de entrada:", data.get('tipo_entrada', '')),
            ("Almacén:", data.get('almacen_nombre', '')),
            ("Fuente de financiamiento:", data.get('fuente_financiamiento_nombre', ''))
        ]

        for i, (label, value) in enumerate(header_data, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws.merge_cells(f'B{i}:F{i}')

        headers = ["Medicamento", "Lote", "Presentación", "Cantidad", "Precio Unitario", "Total"]
        ws.append([''] * 6)
        start_row = ws.max_row + 1
        ws.append(headers)

        for col in range(1, 7):
            cell = ws.cell(row=start_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for item in data.get('items', []):
            row = [
                item.get('nombre', ''), item.get('lote', ''), item.get('presentacion', ''),
                item.get('cantidad', 0), item.get('precio_unitario', 0), item.get('total', 0)
            ]
            ws.append(row)

        column_widths = {'A': 40, 'B': 20, 'C': 25, 'D': 15, 'E': 20, 'F': 20}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=start_row + 1, max_col=6, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            row[4].number_format = '"$"#,##0.00'
            row[5].number_format = '"$"#,##0.00'
            row[3].alignment = Alignment(horizontal='center')

        total_row = ws.max_row + 1
        ws.merge_cells(f'A{total_row}:E{total_row}')
        ws[f'A{total_row}'] = "TOTAL GENERAL:"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'F{total_row}'] = float(data.get('total', 0))
        ws[f'F{total_row}'].font = Font(bold=True)
        ws[f'F{total_row}'].number_format = '"$"#,##0.00'

        firma_row = total_row + 3
        firmas = [('B', "RECIBIDO POR:"), ('D', "AUTORIZADO POR:"), ('F', "ENTREGADO POR:")]
        for col, texto in firmas:
            ws[f'{col}{firma_row}'] = texto
            ws[f'{col}{firma_row + 1}'] = '________________________'
            ws[f'{col}{firma_row + 2}'] = 'Nombre y Firma'
            for offset in [0, 1, 2]:
                ws[f'{col}{firma_row + offset}'].alignment = Alignment(horizontal='center')

        ws[f'A{firma_row + 4}'] = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} por {request.user.get_full_name()}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ENTRADA_{data.get("folio", "REPORTE")}.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
