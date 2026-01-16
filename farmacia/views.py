import os
import json
import openpyxl
import traceback
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db import transaction
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from enfermeria.models import Colectivo, ColectivoMedicamento
from .models import Lote, Medicamento, Presentacion, Proveedor, Entrada, Almacen, DetalleEntrada, Institucion, FuenteFinanciamiento, CPMMedicamento, Receta, RecetaMedicamento, Paciente, MedicamentoNoSurtido
from datetime import timedelta, date, datetime
from math import ceil
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.db.models import Sum, Q, F, Value, IntegerField, Count, Max, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce, TruncMonth
from .forms import LoteForm, MedicamentoForm, SalidaForm
from django.http import HttpResponse, JsonResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import UserSerializer, LoginSerializer
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from django.utils import timezone
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepInFrame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.templatetags.static import static
from .pdf_utils import generar_pdf_salida
from django.conf import settings
from .forms import CargaMasivaForm
from decimal import Decimal
import uuid
from .decorators import group_required, permission_required_or_superuser
import logging
from django.contrib.auth.models import Group, Permission
from django.contrib.auth import get_user_model
logger = logging.getLogger(__name__)
User = get_user_model()



# Create your views here.
def inicio(request):
    return render(request, 'inicio.html')

def vista_farmacia(request):
    return render(request, 'farmacia.html')

def vista_farmacia_g(request):
    """Vista para el inventario por lotes"""
    return render(request, 'farmacia_g.html', {
        'user': request.user
    })

@never_cache
@require_http_methods(["GET", "POST"])
def login_view(request):
    """Vista de login con validación estricta"""
    
    # Si ya está autenticado, redirigir
    if request.user.is_authenticated:
        return redirect('principal')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # Validación básica
        if not username or not password:
            messages.error(request, 'Usuario y contraseña son requeridos')
            return render(request, 'inicio.html', {'username': username})
        
        # ✅ AUTENTICACIÓN ESTRICTA
        user = authenticate(request, username=username, password=password)
        
        # DEBUG: Eliminar en producción
        print(f"🔍 Usuario: {username}")
        print(f"🔍 Authenticate result: {user}")
        print(f"🔍 User is_active: {user.is_active if user else 'N/A'}")
        
        if user is not None:
            if user.is_active:
                # ✅ Limpiar sesión anterior si existe
                request.session.flush()
                
                # ✅ Crear nueva sesión
                login(request, user)
                
                # ✅ Forzar guardado de sesión
                request.session.save()
                
                # Registrar login (opcional)
                print(f"✅ Login exitoso: {user.username} - Rol: {user.rol}")
                
                # Redirigir
                next_url = request.POST.get('next') or request.GET.get('next', 'principal')
                return redirect(next_url)
            else:
                messages.error(request, 'Tu cuenta ha sido desactivada. Contacta al administrador.')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
            print(f"❌ Login fallido para: {username}")
        
        return render(request, 'inicio.html', {'username': username})
    
    # GET request
    return render(request, 'inicio.html')


@never_cache
@login_required
def bienvenida(request):
    def tiene_acceso(user, grupos_requeridos):
        # Acceso para superusuarios O usuarios con rol ADMIN
        if user.is_superuser or user.rol == 'ADMIN':
            return True
        # Verificación por grupos para otros usuarios
        return user.groups.filter(name__in=grupos_requeridos).exists()
    
    modulos = [
        {
            'nombre': 'Farmacia',
            'imagen': 'farmacia/img/farmacia.png',
            'descripcion': 'Gestión de medicamentos y lotes',
            'url': 'farmacia_g',
            'acceso': tiene_acceso(request.user, ['Capturista_Farmacia', 'Supervisor_Farmacia'])
        },
        {
            'nombre': 'Enfermería',
            'imagen': 'farmacia/img/enfermeria.png',
            'descripcion': 'Gestión de pacientes y tratamientos',
            'url': None,
            'acceso': tiene_acceso(request.user, ['Enfermeria'])
        },
        # ... otros módulos con la misma estructura
    ]
    
    return render(request, 'principal.html', {
        'modulos': modulos,
        'last_login': request.user.last_login
    })

@never_cache  # ✅ Evita que el navegador cachee la página
@require_http_methods(["GET", "POST"])  # ✅ Solo permite GET/POST
def logout_view(request):
    """Cierra sesión y destruye completamente la sesión del usuario"""
    if request.user.is_authenticated:
        # Limpia TODA la información de la sesión
        request.session.flush()  # ✅ Destruye la sesión en DB y cookie
        logout(request)  # ✅ Cierra la sesión del usuario
        
        messages.success(request, 'Sesión cerrada correctamente')
    
    # Respuesta con headers de no-cache
    response = redirect('login')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

@login_required
def alertas(request):
    usuario = request.user
    es_admin = usuario.is_superuser or usuario.groups.filter(name='Administrador').exists()
    es_capturista = usuario.groups.filter(name='Capturista').exists()

    # Filtros
    medicamento_id = request.GET.get('medicamento')
    color_filtro = request.GET.get('color')

    lotes = Lote.objects.all().select_related('medicamento')

    if medicamento_id:
        lotes = lotes.filter(medicamento_id=medicamento_id)

    if color_filtro:
        hoy = timezone.now().date()
        if color_filtro == 'verde':
            lotes = lotes.filter(fecha_caducidad__gt=hoy + timedelta(days=365))
        elif color_filtro == 'amarillo':
            lotes = lotes.filter(fecha_caducidad__gt=hoy + timedelta(days=180), fecha_caducidad__lte=hoy + timedelta(days=365))
        elif color_filtro == 'rojo':
            lotes = lotes.filter(fecha_caducidad__lte=hoy + timedelta(days=180))

    medicamentos = Medicamento.objects.all()

    context = {
        'lotes': lotes,
        'medicamentos': medicamentos,
        'es_admin': es_admin,
        'es_capturista': es_capturista,
    }
    return render(request, 'alertas.html', context)



@login_required
@require_http_methods(["POST", "GET"])
def editar_lote(request, lote_id):
    es_admin = (
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrador', 'Administradores']).exists()
    )

    if not es_admin:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)

    lote = get_object_or_404(Lote, id=lote_id)

    if request.method == 'POST':
        try:
            # Campos que todos pueden editar (pero ya filtraste: solo admins llegan aquí)
            cpm = request.POST.get('cpm')
            presentacion_id = request.POST.get('presentacion')

            if cpm:
                lote.cpm = float(cpm)

            if presentacion_id:
                presentacion = get_object_or_404(Presentacion, id=presentacion_id)
                lote.presentacion = presentacion

            # Campos admin (ahora admin = superuser o grupo Administrador/Administradores)
            lote_codigo = request.POST.get('lote_codigo')
            existencia = request.POST.get('existencia')
            fecha_caducidad = request.POST.get('fecha_caducidad')

            if lote_codigo:
                lote.lote_codigo = lote_codigo

            if existencia is not None and existencia != '':
                lote.existencia = int(existencia)

            if fecha_caducidad:
                lote.fecha_caducidad = fecha_caducidad

            lote.save()

            return JsonResponse({'success': True, 'mensaje': 'Lote actualizado correctamente'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    # GET
    return JsonResponse({
        'id': lote.id,
        'medicamento': f"{lote.medicamento.clave} - {lote.medicamento.descripcion}",
        'cpm': str(lote.cpm),
        'presentacion_id': lote.presentacion.id if lote.presentacion else '',
        'lote_codigo': lote.lote_codigo,
        'existencia': lote.existencia,
        'fecha_caducidad': lote.fecha_caducidad.strftime('%Y-%m-%d')
    })


def tiene_acceso_farmacia(user):
    return (
        user.is_authenticated and (
            user.is_superuser
            or user.rol in ['ADMIN', 'FARMACIA']
            or user.groups.filter(name__in=[
                'Administradores',       # el que sí existe
                'Administrador',
                'Capturista_Farmacia',
                'Supervisor_Farmacia'
            ]).exists()
        )
    )


@never_cache
@login_required(login_url='login')
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def farmacia_g(request):
    if request.method == 'POST':
        medicamento_id = request.POST.get('medicamento')
        lote_codigo = request.POST.get('lote_codigo')
        existencia = request.POST.get('existencia')
        presentacion_id = request.POST.get('presentacion')
        nueva_descripcion = request.POST.get('descripcion')

        if medicamento_id and nueva_descripcion:
            medicamento = Medicamento.objects.get(id=medicamento_id)
            medicamento.descripcion = nueva_descripcion
            medicamento.save()
            return JsonResponse({'status': 'success'})

        if medicamento_id and lote_codigo and existencia and presentacion_id:
            medicamento = Medicamento.objects.get(id=medicamento_id)
            presentacion = Presentacion.objects.get(id=presentacion_id)
            
            import uuid
            lote_id = str(uuid.uuid4())[:15]
            
            Lote.objects.create(
                id=lote_id,
                medicamento=medicamento,
                lote_codigo=lote_codigo,
                existencia=int(existencia),
                presentacion=presentacion,
                fecha_caducidad=date.today() + timedelta(days=365),
                cpm=0
            )
            
            return redirect('farmacia_g')

    # Obtener todos los lotes
    lotes = (
        Lote.objects
        .select_related('medicamento', 'presentacion')
        .annotate(
            costo_total=ExpressionWrapper(
                F('existencia') * F('costo_unitario'),
                output_field=DecimalField(max_digits=14, decimal_places=2)
            )
        )
        .order_by('fecha_caducidad')
    )
    
    presentaciones = Presentacion.objects.all()
    
    # Calcular días para caducidad y contadores
    hoy = date.today()
    lotes_con_dias = []
    
    # Contadores para estadísticas (basados en meses)
    vigentes = 0      # > 365 días (más de 1 año)
    por_vencer = 0    # 180-365 días (6 meses a 1 año)
    criticos = 0      # < 180 días (menos de 6 meses)
    
    for lote in lotes:
        dias = (lote.fecha_caducidad - hoy).days
        lote.dias_para_caducidad = dias
        lotes_con_dias.append(lote)
        
        # Clasificar para contadores según tus reglas
        if dias > 365:
            vigentes += 1
        elif dias >= 180:
            por_vencer += 1
        else:
            criticos += 1
    
    medicamentos = Medicamento.objects.filter(activo=True)
    
    es_admin = (
    request.user.is_superuser or
    request.user.groups.filter(name__in=['Administrador', 'Administradores']).exists())

    context = {
        'lotes': lotes_con_dias,
        'medicamentos': medicamentos,
        'presentaciones': presentaciones,
        'vigentes': vigentes,
        'por_vencer': por_vencer,
        'criticos': criticos,
        'es_admin': es_admin,
    }

    return render(request, 'farmacia_g.html', context)


@csrf_exempt
@login_required
@permission_required('farmacia.change_medicamento', raise_exception=True)
def guardar_descripcion(request):
    if request.method == 'POST':
        medicamento_id = request.POST.get('medicamento_id')
        descripcion = request.POST.get('descripcion')
        
        if medicamento_id and descripcion:
            try:
                medicamento = Medicamento.objects.get(id=medicamento_id)
                medicamento.descripcion = descripcion
                medicamento.save()
                return JsonResponse({'status': 'success'})
            except Medicamento.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Medicamento no encontrado'})
    
    return JsonResponse({'status': 'error', 'message': 'Datos inválidos'})


@login_required(login_url='login')
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia', 'Enfermero', 'Jefe de Enfermería', 'Médico')
def inventario_general(request):
    """Vista de inventario general - suma de existencias por medicamento"""
    busqueda = request.GET.get('busqueda', '').strip()
    
    # Obtener inventario agrupado por medicamento
    inventario = Lote.objects.values(
        'medicamento__id',
        'medicamento__clave',
        'medicamento__descripcion',
    ).annotate(
        existencia_total=Sum('existencia'),
        cpm_medicamento=Coalesce(
            F('medicamento__cpm_medicamento__valor'),
            Value(0),
            output_field=IntegerField()
        )
    ).filter(
        existencia_total__gt=0
    ).order_by('medicamento__descripcion')
    
    # Aplicar filtro de búsqueda si existe
    if busqueda:
        inventario = inventario.filter(
            Q(medicamento__descripcion__icontains=busqueda) |
            Q(medicamento__clave__icontains=busqueda)
        )
    
    # ✅ CALCULAR ESTADÍSTICAS DE STOCK
    sobreabasto = 0     # > 100% CPM
    stock_adecuado = 0  # 50-100% CPM
    stock_bajo = 0      # 1-49% CPM
    desabasto = 0       # 0% CPM
    
    # Agregar porcentaje a cada item
    inventario_con_porcentaje = []
    for item in inventario:
        existencia = item['existencia_total']
        cpm = item['cpm_medicamento']
        
        # Calcular porcentaje
        if cpm > 0:
            porcentaje = round((existencia / cpm) * 100, 1)
        else:
            porcentaje = 0  # Sin CPM definido
        
        # Clasificar estado
        if cpm > 0:
            if porcentaje > 100:
                estado = 'sobreabasto'
                sobreabasto += 1
            elif porcentaje >= 50:
                estado = 'adecuado'
                stock_adecuado += 1
            elif porcentaje > 0:
                estado = 'bajo'
                stock_bajo += 1
            else:
                estado = 'desabasto'  # coincide con el <select>
                desabasto += 1
        else:
            estado = 'sin-cpm'
        
        # Agregar datos calculados
        item['porcentaje'] = porcentaje
        item['estado'] = estado
        inventario_con_porcentaje.append(item)
    
    context = {
        'inventario': inventario_con_porcentaje,
        'busqueda_actual': busqueda,
        'total_medicamentos': len(inventario_con_porcentaje),
        'sobreabasto': sobreabasto,  # si quieres mostrarlo en stats
        'stock_adecuado': stock_adecuado,
        'stock_bajo': stock_bajo,
        'desabasto': desabasto,
    }
    
    return render(request, 'inv_gene_f.html', context)



@require_http_methods(['POST'])
@csrf_exempt
@login_required
@permission_required('farmacia.change_cpmmedicamento', raise_exception=True)
def editar_cpm_medicamento(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            medicamento_id = data.get('medicamento_id')
            nuevo_cpm = data.get('cpm')
            
            if not medicamento_id or nuevo_cpm is None:
                return JsonResponse({'error': 'Datos incompletos'}, status=400)
            
            try:
                nuevo_cpm = int(nuevo_cpm)
                if nuevo_cpm < 0:
                    return JsonResponse({'error': 'El CPM no puede ser negativo'}, status=400)
            except ValueError:
                return JsonResponse({'error': 'El CPM debe ser un número válido'}, status=400)
            
            # Obtener o crear el registro CPM
            medicamento = Medicamento.objects.get(id=medicamento_id)
            cpm_obj, created = CPMMedicamento.objects.get_or_create(
                medicamento=medicamento,
                defaults={'valor': nuevo_cpm, 'actualizado_por': request.user}
            )
            
            if not created:
                cpm_obj.valor = nuevo_cpm
                cpm_obj.actualizado_por = request.user
                cpm_obj.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'CPM actualizado correctamente',
                'nuevo_cpm': nuevo_cpm
            })
            
        except Medicamento.DoesNotExist:
            return JsonResponse({'error': 'Medicamento no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
@require_http_methods(['DELETE', 'POST'])
@permission_required('farmacia.delete_lote', raise_exception=True)
def eliminar_lote(request, lote_id):
    """
    Elimina un lote del inventario
    Solo permite eliminar si existencia = 0
    """
    try:
        lote = get_object_or_404(Lote, id=lote_id)
        
        # Información del lote
        medicamento_clave = lote.medicamento.clave
        medicamento_descripcion = lote.medicamento.descripcion
        lote_codigo = lote.lote_codigo
        existencia = lote.existencia
        
        # VALIDACIÓN: Verificar que existencia sea 0
        if existencia > 0:
            return JsonResponse({
                'success': False,
                'tipo': 'error_existencia',
                'error': f'No se puede eliminar el lote {lote_codigo}',
                'detalle': f'El lote tiene {existencia} unidades en existencia.',
                'solucion': 'Para eliminar este lote, primero debes registrar salidas hasta que la existencia sea 0.',
                'existencia': existencia
            }, status=400)
        
        # Si llegamos aquí, existencia = 0, proceder a eliminar
        lote.delete()
        
        return JsonResponse({
            'success': True,
            'mensaje': f'Lote {lote_codigo} eliminado correctamente',
            'clave': medicamento_clave,
            'descripcion': medicamento_descripcion
        })
        
    except Lote.DoesNotExist:
        return JsonResponse({
            'success': False,
            'tipo': 'error_no_existe',
            'error': 'El lote no existe',
            'detalle': 'Es posible que ya haya sido eliminado por otro usuario.'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'tipo': 'error_sistema',
            'error': 'Error del sistema',
            'detalle': str(e)
        }, status=500)


@never_cache
@login_required(login_url='login')
@require_http_methods(['GET', 'POST'])
@permission_required('farmacia.add_medicamento', raise_exception=True)
def registro_medicamento(request):
    """Vista para registrar un nuevo medicamento - Solo clave y descripción"""
    if request.method == 'POST':
        form = MedicamentoForm(request.POST)
        if form.is_valid():
            try:
                medicamento = form.save(commit=False)

                # NO generar medicamento.id: ahora es BigAutoField (numérico)
                medicamento.activo = True
                medicamento.costo = 0.00
                medicamento.codigo_barras = None
                medicamento.proveedor = None
                medicamento.presentacion = None

                medicamento.save()

                messages.success(
                    request,
                    f'✓ Medicamento "{medicamento.clave}" registrado correctamente. '
                    f'Ahora puedes agregar lotes desde el módulo de Entradas.'
                )
                return redirect('farmacia_g')

            except Exception as e:
                messages.error(request, f'Error al registrar medicamento: {str(e)}')
        else:
            messages.error(request, 'Error en el formulario. Verifica los datos.')
    else:
        form = MedicamentoForm()

    return render(request, 'registro_medicamento.html', {'form': form})



@never_cache
@login_required(login_url='login')
@permission_required('farmacia.add_entrada', raise_exception=True)
def entrada_medicamentos(request): 
    context = {
        'presentaciones': Presentacion.objects.all(),
        'almacenes': Almacen.objects.all(),
        'instituciones': Institucion.objects.all(),
        'fuentes_financiamiento': FuenteFinanciamiento.objects.all(),
        'hoy': date.today().isoformat(),
        'user': request.user
    }

    # Lógica para folio (GET o POST fallido)
    if 'folio_entrada' in request.GET:  # Para previsualización
        context['folio_entrada'] = request.GET.get('folio_entrada')
    else:
        # Generar folio automático
        date_str = date.today().strftime('%Y%m%d')
        last_entry = Entrada.objects.filter(folio__startswith=f'ENT-{date_str}').order_by('-folio').first()
        new_num = int(last_entry.folio.split('-')[-1]) + 1 if last_entry else 1
        context['folio_entrada'] = f"ENT-{date_str}-{new_num:04d}"

    # Manejo de POST (guardar entrada)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Obtener folio (manual o automático)
                folio = request.POST.get('folio_entrada')
                if not folio:  # Si no se proporcionó folio manual
                    folio = context['folio_entrada']  # Usar el generado automáticamente

                # Validación básica
                if not folio:
                    raise ValueError("El folio es requerido")

                # Crear la entrada (ajusta según tu modelo)
                entrada = Entrada(
                    folio=folio,
                    medicamento_id=request.POST.get('medicamento_id'),
                    lote=request.POST.get('lote'),
                    caducidad=request.POST.get('caducidad'),
                    cantidad=request.POST.get('cantidad'),
                    # ... otros campos ...
                )
                entrada.save()

                messages.success(request, f'Entrada {folio} guardada correctamente')
                return redirect('nombre_de_tu_url_de_exito')  # Ajusta la URL

        except Exception as e:
            messages.error(request, f'Error al guardar: {str(e)}')
            # Mantenemos los valores en el contexto para repoblar el formulario
            context.update({
                'folio_entrada': request.POST.get('folio_entrada', context['folio_entrada']),
                # ... otros campos del formulario ...
            })

    return render(request, 'entrada_medicamentos.html', context)


# En views.py - agregar esta vista
@csrf_exempt
@login_required
def buscar_medicamentos_autocomplete(request):
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse([], safe=False)
    
    medicamentos = Medicamento.objects.filter(
        Q(descripcion__icontains=query) | Q(clave__icontains=query),
        activo=True
    ).select_related('presentacion')[:10]
    
    resultados = []
    for med in medicamentos:
        resultados.append({
            'id': med.id,
            'clave': med.clave,
            'descripcion': med.descripcion,
            'presentacion': med.presentacion.nombre if med.presentacion else 'UNIDAD'
        })
    
    return JsonResponse(resultados, safe=False)

@login_required
def buscar_medicamentos(request):
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse([], safe=False)
    
    medicamentos = Medicamento.objects.filter(
        Q(clave__icontains=query) | Q(descripcion__icontains=query),
        activo=True
    ).select_related('presentacion')[:10]
    
    resultados = []
    for med in medicamentos:
        resultados.append({
            'id': med.id,
            'clave': med.clave,
            'descripcion': med.descripcion,
            'presentacion': med.presentacion.nombre if med.presentacion else 'UNIDAD'
        })
    
    return JsonResponse(resultados, safe=False)

@csrf_exempt
@login_required
def guardar_entradas(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)

        # Campos requeridos (snake_case)
        required_fields = [
            'folio', 'fecha', 'tipo_entrada', 'recibido_por',
            'detalles', 'fuente_financiamiento', 'proceso'
        ]
        for field in required_fields:
            if field not in data or data[field] in (None, '', []):
                return JsonResponse({'error': f'Campo {field} es requerido'}, status=400)

        if not isinstance(data['detalles'], list) or len(data['detalles']) == 0:
            return JsonResponse({'error': 'Debe incluir al menos un medicamento'}, status=400)

        tipo = data['tipo_entrada']
        almacen_id = data.get('almacen')
        institucion_id = data.get('institucion')

        # Validar origen según tipo
        if tipo == 'ALMACEN':
            if not almacen_id:
                return JsonResponse({'error': 'Seleccione un almacén'}, status=400)
            institucion_id = None
        elif tipo == 'TRANSFERENCIA':
            if not institucion_id:
                return JsonResponse({'error': 'Seleccione una institución'}, status=400)
            almacen_id = None
        else:
            return JsonResponse({'error': 'Tipo de entrada inválido'}, status=400)

        with transaction.atomic():
            entrada = Entrada.objects.create(
                folio=data['folio'],
                fecha=data['fecha'],
                tipo_entrada=tipo,
                almacen_id=almacen_id,
                institucion_id=institucion_id,
                fuente_financiamiento_id=data['fuente_financiamiento'],
                contrato=data.get('contrato', ''),
                proceso=data['proceso'],
                recibido_por_id=data['recibido_por'],
                observaciones=data.get('observaciones', '')
            )

            detalle_fields = {f.name for f in DetalleEntrada._meta.fields}

            for det in data['detalles']:
                detalle_required = [
                    'medicamento_id', 'lote', 'caducidad',
                    'cantidad', 'precio_unitario', 'presentacion_id'
                ]
                for f in detalle_required:
                    if f not in det or det[f] in (None, ''):
                        return JsonResponse({'error': f'Campo {f} es requerido en los detalles'}, status=400)

                # Crear detalle con compatibilidad por si el campo se llama distinto
                kwargs_det = dict(
                    entrada=entrada,
                    medicamento_id=det['medicamento_id'],
                    lote=det['lote'],
                    caducidad=det['caducidad'],
                    cantidad=det['cantidad'],
                    presentacion_id=det['presentacion_id'],
                )
                if 'precio_unitario' in detalle_fields:
                    kwargs_det['precio_unitario'] = det['precio_unitario']
                elif 'preciounitario' in detalle_fields:
                    kwargs_det['preciounitario'] = det['precio_unitario']
                else:
                    return JsonResponse({'error': 'El modelo DetalleEntrada no tiene campo de precio unitario'}, status=500)

                DetalleEntrada.objects.create(**kwargs_det)

            return JsonResponse({
                'success': True,
                'folio': entrada.folio,
                'redirect_url': reverse('farmacia_g')
            })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)



@csrf_exempt
@login_required
def generar_reporte_pdf(request):    
    try:
        data = json.loads(request.body)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                              rightMargin=20*mm, leftMargin=20*mm,
                              topMargin=15*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        elements = []
        
        # 1. Definir estilo para medicamentos primero
        medicamento_style = ParagraphStyle(
            name='MedicamentoStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=9,
            alignment=0,
            wordWrap='LTR',
            splitLongWords=True
        )
        
        # 2. Agregar logo
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia/static/farmacia/img/logo.jpg')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=180*mm, height=(175*180/1236)*mm)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 15*mm))
        
        # 3. Título
        titulo = Paragraph('<para align=center><font size=14><b>REPORTE DE ENTRADA DE MEDICAMENTOS</b></font></para>', styles['Normal'])
        elements.append(titulo)
        elements.append(Spacer(1, 8*mm))
        
        # 4. Información de cabecera
        info_data = [
            ['Folio: ', data.get('folio', 'N/A'), 'Fecha: ', data.get('fecha', 'N/A')],
            ['Tipo Entrada: ', data.get('tipo_entrada', 'N/A'), 'Almacén: ', data.get('almacen_nombre', 'N/A')],
            ['Fuente Financ.: ', data.get('fuente_financiamiento_nombre', 'N/A'), 'Proceso: ', data.get('proceso', 'N/A')]
        ]
        
        info_table = Table(info_data, colWidths=[35*mm, 60*mm, 40*mm, 60*mm])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 8*mm))
        
        # 5. Tabla de items
        encabezados = [
            Paragraph('<b>Medicamento</b>', styles['Normal']),
            Paragraph('<b>Lote</b>', styles['Normal']),
            Paragraph('<b>Presentación</b>', styles['Normal']),
            Paragraph('<b>Cantidad</b>', styles['Normal']),
            Paragraph('<b>P. Unitario</b>', styles['Normal']),
            Paragraph('<b>Total</b>', styles['Normal'])
        ]
        datos_tabla = [encabezados]

        # Procesar items una sola vez
        for item in data.get('items', []):
            fila = [
                Paragraph(item.get('nombre', '')),  # Usa el estilo por defecto
                item.get('lote', ''),
                item.get('presentacion', ''),
                str(item.get('cantidad', 0)),
                f"${float(item.get('precio_unitario', 0)):,.2f}",
                f"${float(item.get('total', 0)):,.2f}"
                ]
            datos_tabla.append(fila)

        # Total general (fuera del bucle)
        datos_tabla.append([
            '', '', '', '',
            Paragraph('<b>TOTAL GENERAL:</b>', styles['Normal']),
            f"${float(data.get('total', 0)):,.2f}"
        ])
        
        # Crear tabla con ajustes
        tabla = Table(datos_tabla, colWidths=[80*mm, 25*mm, 35*mm, 20*mm, 25*mm, 25*mm])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (4,-1), (-1,-1), colors.HexColor('#70AD47')),
            ('TEXTCOLOR', (4,-1), (-1,-1), colors.whitesmoke),
            ('FONTNAME', (4,-1), (-1,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('FONTSIZE', (0,1), (-1,-2), 8),
            ('LEADING', (0,1), (0,-2), 9),
        ]))
        
        elements.append(tabla)
        elements.append(Spacer(1, 15*mm))
        
        # 6. Firmas
        firmas_data = [
            ['', '', ''],
            ['________________________', '________________________', '________________________'],
            ['Recibido por', 'Autorizado por', 'Entregado por'],
            ['Nombre y Firma', 'Nombre y Firma', 'Nombre y Firma']
        ]
        
        firmas_table = Table(firmas_data, colWidths=[60*mm, 60*mm, 60*mm])
        firmas_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('SPAN', (0,0), (2,0)),
            ('LEADING', (1,1), (2,1), 14),
        ]))
        elements.append(firmas_table)
        
        # 7. Pie de página
        elements.append(Spacer(1, 10*mm))
        footer = Paragraph(
            f"<font size=7>Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} por {request.user.get_full_name()} | Sistema de Gestión Farmacéutica</font>", 
            styles['Normal'])
        elements.append(footer)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="ENTRADA_{data.get("folio", "REPORTE")}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Error al generar PDF: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)
                

@csrf_exempt
def generar_reporte_excel(request):
    try:
        data = json.loads(request.body)
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Entradas"

        # ===== 1. CONFIGURACIÓN INICIAL =====
        # Estilo de borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ===== 2. LOGO (1236x175 px) =====
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia', 'static', 'farmacia', 'img', 'logo.jpg')
        if os.path.exists(logo_path):
            img = ExcelImage(logo_path)
            # Ajuste preciso de dimensiones (1236x175 px → Excel usa puntos, 1px ≈ 0.75pt)
            img.width = 1236 * 0.75  # Ancho en puntos
            img.height = 175 * 0.75  # Alto en puntos
            ws.add_image(img, 'A1')  # Logo en celda A1
            ws.row_dimensions[1].height = 135  # 175px ≈ 135 puntos (ajuste empírico)

        # ===== 3. CABECERA =====
        # Título principal
        ws.merge_cells('A3:F3')
        title_cell = ws['A3']
        title_cell.value = "REPORTE DE ENTRADA DE MEDICAMENTOS"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Datos de cabecera
        header_data = [
            ("Folio:", data.get('folio', '')),
            ("Fecha:", data.get('fecha', '')),
            ("Tipo de entrada:", data.get('tipo_entrada', '')),
            ("Almacén:", data.get('almacen_nombre', '')),
            ("Fuente de financiamiento:", data.get('fuente_financiamiento_nombre', ''))
        ]

        for i, (label, value) in enumerate(header_data, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws.merge_cells(f'B{i}:F{i}')

        # ===== 4. TABLA DE ITEMS =====
        # Encabezados de tabla
        headers = ["Medicamento", "Lote", "Presentación", "Cantidad", "Precio Unitario", "Total"]
        ws.append([''] * 6)  # Espacio antes de la tabla
        start_row = ws.max_row + 1
        ws.append(headers)

        # Estilo para encabezados
        for col in range(1, 7):
            cell = ws.cell(row=start_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # ===== 5. LLENAR DATOS =====
        for item in data.get('items', []):
            row = [
                item.get('nombre', ''),
                item.get('lote', ''),
                item.get('presentacion', ''),
                item.get('cantidad', 0),
                item.get('precio_unitario', 0),
                item.get('total', 0)
            ]
            ws.append(row)

        # ===== 6. FORMATO Y AJUSTES =====
        # Ajuste de columnas (ancho automático + mínimo)
        column_widths = {
            'A': 40,  # Medicamento
            'B': 20,  # Lote
            'C': 25,  # Presentación
            'D': 15,  # Cantidad
            'E': 20,  # Precio
            'F': 20   # Total
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Formato numérico y alineación
        for row in ws.iter_rows(min_row=start_row + 1, max_col=6, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Formato moneda para precios
            row[4].number_format = '"$"#,##0.00'
            row[5].number_format = '"$"#,##0.00'
            row[3].alignment = Alignment(horizontal='center')  # Cantidad centrada

        # ===== 7. TOTAL GENERAL =====
        total_row = ws.max_row + 1
        ws.merge_cells(f'A{total_row}:E{total_row}')
        ws[f'A{total_row}'] = "TOTAL GENERAL:"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'F{total_row}'] = float(data.get('total', 0))
        ws[f'F{total_row}'].font = Font(bold=True)
        ws[f'F{total_row}'].number_format = '"$"#,##0.00'

        # ===== 8. FIRMAS =====
        firma_row = total_row + 3
        firmas = [
            ('B', "RECIBIDO POR:"),
            ('D', "AUTORIZADO POR:"),
            ('F', "ENTREGADO POR:")
        ]

        for col, texto in firmas:
            ws[f'{col}{firma_row}'] = texto
            ws[f'{col}{firma_row + 1}'] = '________________________'
            ws[f'{col}{firma_row + 2}'] = 'Nombre y Firma'
            
            for offset in [0, 1, 2]:
                ws[f'{col}{firma_row + offset}'].alignment = Alignment(horizontal='center')

        # ===== 9. PIE DE PÁGINA =====
        ws[f'A{firma_row + 4}'] = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} por {request.user.get_full_name()}"

        # ===== 10. GUARDAR =====
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ENTRADA_{data.get("folio", "REPORTE")}.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


class RegisterAPIView(APIView):
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            refresh = RefreshToken.for_user(user)
            return Response({
                'user': serializer.data,
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LoginAPIView(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            })
        return Response(serializer.errors, status=status.HTTP_401_UNAUTHORIZED)
    

@never_cache
@login_required(login_url='login')
@permission_required('farmacia.create_salida', raise_exception=True)
def registrar_salida(request):
    if request.method == 'POST':
        
        # 1. Extraer datos del paciente
        curp = request.POST.get('paciente_curp', '').strip().upper()
        nombre = request.POST.get('paciente_nombre')
        nacimiento_str = request.POST.get('paciente_nacimiento')
        origen = request.POST.get('receta_origen')
        folio = request.POST.get('receta_folio')
        
        # 2. Extraer items surtidos
        items_para_guardar = []
        index = 0
        while True:
            lote_id = request.POST.get(f'item_lote_{index}')
            cantidad_str = request.POST.get(f'item_cantidad_{index}')
            if not lote_id or not cantidad_str: 
                break
            try:
                lote = Lote.objects.get(id=lote_id)
                cantidad = int(cantidad_str)
                if cantidad <= 0: 
                    raise Exception(f"Cantidad inválida para {lote.lote_codigo}")
                if cantidad > lote.existencia: 
                    raise Exception(f"Stock insuficiente para {lote.lote_codigo}")
                items_para_guardar.append({'lote': lote, 'cantidad': cantidad})
                index += 1
            except (Lote.DoesNotExist, ValueError, Exception) as e:
                return JsonResponse({"success": False, "error": str(e)}, status=400)
        
        # 3. ✅ EXTRAER MEDICAMENTOS FALTANTES (NUEVO)
        medicamentos_faltantes = []
        index = 0
        while True:
            faltante_desc = request.POST.get(f'faltante_desc_{index}')
            faltante_cant_str = request.POST.get(f'faltante_cant_{index}')
            faltante_motivo = request.POST.get(f'faltante_motivo_{index}')
            
            if not faltante_desc or not faltante_cant_str or not faltante_motivo:
                break
            
            try:
                medicamentos_faltantes.append({
                    'descripcion': faltante_desc,
                    'cantidad': int(faltante_cant_str),
                    'motivo': faltante_motivo
                })
                index += 1
            except ValueError as e:
                return JsonResponse({
                    "success": False, 
                    "error": f"Cantidad inválida para medicamento faltante: {str(e)}"
                }, status=400)
        
        # 4. Validar que haya al menos algo que registrar
        if not items_para_guardar and not medicamentos_faltantes:
            return JsonResponse({
                "success": False, 
                "error": "No hay medicamentos en la lista ni medicamentos faltantes registrados."
            }, status=400)
        
        # 5. Guardar todo
        try:
            with transaction.atomic():
                
                # 5.1. Validar fecha de nacimiento
                try:
                    nacimiento_obj = datetime.strptime(nacimiento_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    return JsonResponse({
                        "success": False, 
                        "error": "Fecha de nacimiento inválida."
                    }, status=400)
                
                # 5.2. Buscar o Crear Paciente
                if curp:
                    # Si SÍ hay CURP, lo usamos como llave
                    paciente, _ = Paciente.objects.update_or_create(
                        curp=curp,
                        defaults={
                            'nombre_completo': nombre, 
                            'fecha_nacimiento': nacimiento_obj
                        }
                    )
                else:
                    # Si NO hay CURP, usamos Nombre + Fecha Nacimiento como llave
                    paciente, _ = Paciente.objects.get_or_create(
                        nombre_completo=nombre,
                        fecha_nacimiento=nacimiento_obj,
                        defaults={'curp': None}
                    )
                
                # 5.3. ✅ DETERMINAR ESTADO DE LA RECETA (NUEVO)
                if medicamentos_faltantes and items_para_guardar:
                    estado_receta = 'parcial'  # Algunos surtidos, algunos no
                elif medicamentos_faltantes and not items_para_guardar:
                    estado_receta = 'no_surtida'  # Ninguno disponible
                else:
                    estado_receta = 'completa'  # Todos surtidos
                
                # 5.4. Crear Receta
                if not folio:
                    # Generar folio automático con formato mejorado
                    fecha_str = timezone.now().strftime('%Y%m%d')
                    ultimo = Receta.objects.filter(
                        id_folio__startswith=f'REC-{fecha_str}'
                    ).order_by('-id_folio').first()
                    
                    if ultimo:
                        ultimo_num = int(ultimo.id_folio.split('-')[-1])
                        folio = f"REC-{fecha_str}-{ultimo_num + 1:04d}"
                    else:
                        folio = f"REC-{fecha_str}-0001"
                
                receta_salida = Receta.objects.create(
                    id_folio=folio,
                    paciente=paciente, 
                    fecha_emision=timezone.now().date(),
                    fecha_surtido=timezone.now().date(), 
                    estado=estado_receta,  # ✅ Estado dinámico
                    origen=origen,
                    surtido_por=request.user 
                )
                
                # 5.5. Guardar items surtidos y restar stock
                for item in items_para_guardar:
                    lote = item['lote']
                    cantidad = item['cantidad']
                    
                    precio_unitario = (lote.costo_unitario if lote else Decimal('0.00'))
                    precio_total = Decimal(cantidad) * precio_unitario 
                       
                    RecetaMedicamento.objects.create(
                        receta=receta_salida,
                        medicamento=lote.medicamento, 
                        lote=lote, 
                        cantidad_solicitada=cantidad,
                        cantidad_surtida=cantidad,
                        precio_unitario=precio_unitario,
                        precio_total=precio_total,
                    )
                    
                    # Restar del stock
                    lote.existencia -= cantidad
                    lote.save(update_fields=['existencia'])
                
                # 5.6. ✅ GUARDAR MEDICAMENTOS NO SURTIDOS (NUEVO)
                for faltante in medicamentos_faltantes:
                    MedicamentoNoSurtido.objects.create(
                        receta=receta_salida,
                        medicamento_descripcion=faltante['descripcion'],
                        cantidad_solicitada=faltante['cantidad'],
                        motivo=faltante['motivo'],
                        registrado_por=request.user
                    )
            
            # 6. Éxito: Devolver JSON con información adicional
            pdf_url = reverse('descargar_comprobante', args=[receta_salida.pk])
            
            mensaje_estado = {
                'completa': '✓ Todos los medicamentos fueron surtidos.',
                'parcial': '⚠ Surtido parcial. Algunos medicamentos no estaban disponibles.',
                'no_surtida': '✗ Ningún medicamento pudo ser surtido.'
            }
            
            return JsonResponse({
                "success": True, 
                "message": f"Salida registrada: {folio}",
                "pdf_url": pdf_url,
                "estado": estado_receta,
                "mensaje_estado": mensaje_estado.get(estado_receta, ''),
                "items_surtidos": len(items_para_guardar),
                "items_faltantes": len(medicamentos_faltantes)
            })
        
        except Exception as e:
            import traceback
            traceback.print_exc() 
            return JsonResponse({
                "success": False, 
                "error": str(e)
            }, status=500)
    
    # --- LÓGICA GET ---
    form = SalidaForm() 
    context = {
        'form': form,
        'titulo_pagina': 'Registro de Salidas'
    }
    return render(request, 'salida_medicamentos.html', context)



# ==================================================
# VISTA 2: DESCARGAR EL PDF (NUEVA)
# ==================================================
@login_required
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def descargar_comprobante(request, receta_id):
    try:
        # Buscamos la receta por su folio
        receta = get_object_or_404(Receta.objects.select_related('paciente', 'surtido_por'), pk=receta_id)
        
        # 4. Generar el PDF
        pdf_buffer = generar_pdf_salida(receta) # ¡Función actualizada!
        
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="salida_{receta.id_folio}.pdf"'
        return response
        
    except Exception as e:
        messages.error(request, f"Error al generar el PDF: {e}")
        return redirect('registrar_salida')
    

@login_required
@require_http_methods(['POST'])
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def generar_excel_salidas(request):
    """
    Genera un reporte en Excel de todas las salidas (RecetaMedicamento)
    en un rango de fechas.
    """
    try:
        # Asumo que envías un JSON con 'fecha_inicio' y 'fecha_fin'
        # Si no, necesitarás un formulario
        
        # Si no usas JSON, y usas un form normal, sería request.POST.get('fecha_inicio')
        data = json.loads(request.body) 
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')

        # Si las fechas no vienen, puedes usar un rango por defecto (ej. el día de hoy)
        if not fecha_inicio:
            fecha_inicio = timezone.now().replace(hour=0, minute=0, second=0)
        if not fecha_fin:
            fecha_fin = timezone.now().replace(hour=23, minute=59, second=59)

        # 1. La consulta a la base de datos (la parte clave)
        salidas = RecetaMedicamento.objects.filter(
            receta__fecha_surtido__range=(fecha_inicio, fecha_fin)
        ).order_by('receta__fecha_surtido', 'lote__medicamento__descripcion')

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Salidas"

        # 2. Títulos
        titulos = ['Fecha Surtido', 'Área', 'Médico/Quien Solicita', 'Clave', 'Medicamento', 'Lote', 'Cantidad Surtida']
        ws.append(titulos)
        
        # (Opcional: Ponerlos en negrita)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 3. Datos
        for item in salidas:
            ws.append([
                item.receta.fecha_surtido.strftime('%d/%m/%Y %H:%M'),
                item.receta.area.nombre if item.receta.area else 'N/A',
                item.receta.medico,
                item.lote.medicamento.clave,
                item.lote.medicamento.descripcion,
                item.lote.lote_codigo,
                item.cantidad_surtida
            ])

        # (Aquí puedes ajustar anchos de columna, etc.)

        # 4. Guardar en buffer y devolver
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="REPORTE_SALIDAS_{fecha_inicio}_{fecha_fin}.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
    
def get_paciente_info_json(request, curp):
    """
    Vista que devuelve la información de un Paciente como JSON
    para ser usada por JavaScript al teclear el CURP.
    """
    if request.method == "GET":
        try:
            # --- ¡CAMBIO IMPORTANTE! ---
            # Usamos .get() en lugar de get_object_or_404
            # para poder atrapar el error 'DoesNotExist' nosotros mismos.
            paciente = Paciente.objects.get(curp=curp.upper())
            
            # Si lo encuentra, preparamos los datos
            data = {
                'id': paciente.id,
                'nombre_completo': paciente.nombre_completo,
                'curp': paciente.curp,
                # El input HTML de tipo 'date' necesita el formato AAAA-MM-DD
                'fecha_nacimiento': paciente.fecha_nacimiento.strftime('%Y-%m-%d'), 
            }
            return JsonResponse(data) # Devuelve 200 OK
        
        except Paciente.DoesNotExist: 
            # ¡Esto es lo que queremos! Significa que es un paciente nuevo.
            # El JS está esperando este error 404 para dejarte escribir.
            return JsonResponse({'error': 'Paciente no encontrado'}, status=404)
        except Exception as e:
            # Atrapa cualquier otro error 
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def buscar_lote_json(request, query):
    """
    Vista UNIFICADA que busca un lote.
    Intenta buscar por ID (pk) O por Código de Lote.
    """
    if request.method == "GET":
        try:
            # ¡Aquí está la magia!
            # Buscamos un Lote donde el ID (pk) sea la query
            # O (el | significa 'OR') el lote_codigo sea la query
            lote = Lote.objects.get(
                Q(id=query) | Q(lote_codigo=query.upper())
            )
            
            # Si lo encontramos (por cualquiera de los dos), devolvemos los datos
            data = {
                'id': lote.id, # El ID (pk) que usará el formulario ('auto-B1011')
                'medicamento_nombre': lote.medicamento.descripcion,
                'clave': lote.medicamento.clave,
                'lote_numero': lote.lote_codigo,
                'caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y'),
                'cantidad_actual': lote.existencia,
            }
            return JsonResponse(data)
        
        except Lote.DoesNotExist: 
            return JsonResponse({'error': f'Lote "{query}" no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def get_paciente_by_name(request, nombre):
    """
    Busca al primer paciente que coincida (ignorando mayúsculas) con el nombre.
    """
    if request.method == "GET":
        try:
            # Usamos filter() + first() porque los nombres se pueden repetir
            paciente = Paciente.objects.filter(nombre_completo__iexact=nombre).first()
            
            if paciente: # Si encontró uno
                data = {
                    'id': paciente.id,
                    'nombre_completo': paciente.nombre_completo,
                    'curp': paciente.curp or '', # Devolvemos '' si es None
                    'fecha_nacimiento': paciente.fecha_nacimiento.strftime('%Y-%m-%d'), 
                }
                return JsonResponse(data)
            else:
                # No es un error, solo que no existe
                return JsonResponse({'error': 'Paciente no encontrado'}, status=404)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
@permission_required('farmacia.view_carga_masiva', raise_exception=True)
def carga_masiva(request):
    """Vista para mostrar el formulario de carga masiva"""
    form = CargaMasivaForm()
    return render(request, 'carga_masiva.html', {
        'form': form,
        'user': request.user
    })

@csrf_exempt
@login_required
def procesar_carga_masiva(request):
    """Procesa el archivo Excel de carga masiva"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    form = CargaMasivaForm(request.POST, request.FILES)
    if not form.is_valid():
        errores = []
        for field, errors in form.errors.items():
            for error in errors:
                errores.append(str(error))
        return JsonResponse({'error': ', '.join(errores)}, status=400)
    
    from .utils import ProcesadorCargaMasiva
    archivo = form.cleaned_data['archivo']
    procesador = ProcesadorCargaMasiva(archivo)
    resultado = procesador.procesar()
    
    # Si hay error crítico en el procesamiento
    if 'error' in resultado:
        return JsonResponse(resultado, status=400)
    
    # Preparar respuesta con advertencias
    response_data = {
        'success': True,
        'mensaje': 'Carga masiva completada',
        'resultados': {
            'total': resultado['resultados']['total'],
            'exitosos': resultado['resultados']['exitosos'],
            'actualizados': resultado['resultados']['actualizados'],
            'errores': resultado['resultados']['errores'],
            'advertencias': resultado['resultados'].get('advertencias', [])  # ← NUEVO
        }
    }
    
    # Código de estado basado en si hubo errores
    status_code = 200 if len(resultado['resultados']['errores']) == 0 else 207  # 207 = Multi-Status
    
    return JsonResponse(response_data, status=status_code)

@login_required
def exportar_inventario_excel(request):
    """Exportar inventario a Excel con logo"""
    try:
        import xlsxwriter
        
        lotes = Lote.objects.select_related('medicamento', 'presentacion').all()
        
        # Crear archivo en memoria
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Inventario')
        
        # Definir formatos
        header_format = workbook.add_format({
            'bg_color': '#8B0000',
            'font_color': 'white',
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_size': 11
        })
        
        title_format = workbook.add_format({
            'bg_color': '#8B0000',
            'font_color': 'white',
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 14
        })
        
        date_format = workbook.add_format({
            'italic': True,
            'align': 'left',
            'font_size': 10
        })
        
        data_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_size': 10
        })
        
        red_format = workbook.add_format({
            'bg_color': '#FF0000',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'bold': True
        })
        
        orange_format = workbook.add_format({
            'bg_color': '#FF4444',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        yellow_format = workbook.add_format({
            'bg_color': '#FFFF00',
            'font_color': 'black',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        green_format = workbook.add_format({
            'bg_color': '#00B050',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Establecer ancho de columnas
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 40)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 18)
        worksheet.set_column('E:E', 12)
        worksheet.set_column('F:F', 15)
        worksheet.set_column('F:F', 12)  # Costo
        worksheet.set_column('G:G', 15) 
        
        # Agregar logo
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia', 'static', 'farmacia', 'img', 'logo.jpg')
        if os.path.exists(logo_path):
            try:
                # Insertar logo en A1, redimensionado
                worksheet.insert_image('A1', logo_path, {
                    'x_scale': 0.8,
                    'y_scale': 0.8,
                    'x_offset': 0,
                    'y_offset': 0
                })
            except Exception as e:
                print(f"Error insertando logo: {e}")
        
        # Título (desplazado después del logo)
        worksheet.merge_range('A3:H3', 'REPORTE DE INVENTARIO DE MEDICAMENTOS POR LOTE', title_format)
        worksheet.merge_range('A4:H4', f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_format)

        
        # Encabezados
        headers = ['Clave', 'Descripción', 'Lote', 'Presentación', 'Existencia', 'Costo Unit.','Caducidad']
        for col, header in enumerate(headers):
            worksheet.write(5, col, header, header_format)
        worksheet.set_row(5, 20)
        
        # Datos
        row = 6
        for lote in lotes:
            worksheet.write(row, 0, lote.medicamento.clave, data_format)
            worksheet.write(row, 1, lote.medicamento.descripcion, data_format)
            worksheet.write(row, 2, lote.lote_codigo, data_format)
            worksheet.write(row, 3, lote.presentacion.nombre if lote.presentacion else 'N/A', data_format)
            worksheet.write(row, 4, lote.existencia, data_format)
            worksheet.write(row, 5, float(lote.costo_unitario or 0), data_format)
            
            # Determinar color según caducidad
            dias_restantes = (lote.fecha_caducidad - datetime.now().date()).days
            fecha_str = lote.fecha_caducidad.strftime('%d/%m/%Y')
            
            if dias_restantes <= 0:
                worksheet.write(row, 6, fecha_str, red_format)
            elif dias_restantes <= 30:
                worksheet.write(row, 6, fecha_str, orange_format)
            elif dias_restantes <= 90:
                worksheet.write(row, 6, fecha_str, yellow_format)
            else:
                worksheet.write(row, 6, fecha_str, green_format)
            
            worksheet.set_row(row, 18)
            row += 1
        
        # Pie de página
        worksheet.merge_range(row + 1, 0, row + 1, 7, 'Documento generado automáticamente por INVENTFARM', date_format)
        
        # Finalizar
        workbook.close()
        
        # Preparar respuesta
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Inventario_{datetime.now().strftime("%d%m%Y")}.xlsx"'
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error: {str(e)}', status=400)



@login_required
def exportar_inventario_pdf(request):
    """Exportar inventario a PDF con logo"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Paragraph
        from datetime import datetime  # ✅ Asegúrate de importar
        from reportlab.lib.styles import getSampleStyleSheet

        
        lotes = Lote.objects.select_related('medicamento', 'presentacion').all()
        
        # Crear PDF en memoria
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter  # 8.5" x 11"
        
        # ✅ RUTA CORRECTA (igual a pdf_utils.py)
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia', 'static', 'farmacia', 'img', 'logo.jpg')
        
        # Configurar logo (dimensiones 1236x175 -> ratio ~7:1)
        logo_width = 7.0 * inch
        logo_height = 1.0 * inch
        
        # Centrar logo
        x_logo = (width - logo_width) / 2
        y_logo = height - (0.75 * inch) - logo_height
        
        # Dibujar logo
        if os.path.exists(logo_path):
            try:
                p.drawImage(logo_path, x_logo, y_logo,
                           width=logo_width, height=logo_height,
                           preserveAspectRatio=True)
            except Exception as e:
                print(f"Error cargando logo: {e}")
        else:
            print(f"Logo no encontrado en: {logo_path}")
        
        # Posición para el contenido después del logo
        y_actual = y_logo - (0.25 * inch)
        
        # Título
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width / 2.0, y_actual, "REPORTE DE INVENTARIO POR LOTES DE MEDICAMENTOS")
        y_actual -= 20
        
        # Fecha
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2.0, y_actual, f"Fecha de Generacion: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y_actual -= 16
        
        # ✅ USUARIO QUE GENERÓ EL PDF
        nombre_usuario = (request.user.get_full_name() or request.user.username).strip()
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2.0, y_actual, f"Generado por: {nombre_usuario}")
        y_actual -= 20
        
        # Línea divisoria
        p.line(inch, y_actual, width - inch, y_actual)
        y_actual -= 20
        
        # Tabla de datos
        data_tabla = [['Clave', 'Descripción', 'Lote', 'Pres.', 'Exist.', 'Costo Unit.', 'Caducidad']]
        
        styles = getSampleStyleSheet()
        style_desc = styles["Normal"]
        style_desc.fontName = "Helvetica"
        style_desc.fontSize = 7
        style_desc.leading = 8
        
        style_clave = styles["Normal"]
        style_clave.fontName = "Helvetica"
        style_clave.fontSize = 6.2
        style_clave.leading = 7

        
        for lote in lotes:
            desc = lote.medicamento.descripcion or ""
            clave = lote.medicamento.clave or ""
            data_tabla.append([
                Paragraph(clave, style_clave),
                Paragraph(desc, style_desc),   # ✅ WRAP real
                lote.lote_codigo,
                lote.presentacion.nombre[:8] if lote.presentacion else 'N/A',
                str(lote.existencia),
                f"${lote.costo_unitario:.2f}",
                lote.fecha_caducidad.strftime('%d/%m/%Y')
            ])
        
        # ✅ Ancho disponible real en la hoja (dejando márgenes)
        margen_x = inch
        ancho_disponible = width - (2 * margen_x)

        # Pesos por columna (más peso = más ancho)
        pesos = {
            "clave": 1.5,
            "descripcion": 3.6,   # aquí decides cuánto gana descripción
            "lote": 1.4,
            "pres": 1.0,
            "exist": 0.9,
            "costo": 1.1,
            "cad": 1.3,
        }

        total = sum(pesos.values())
        col_widths = [
            ancho_disponible * (pesos["clave"] / total),
            ancho_disponible * (pesos["descripcion"] / total),
            ancho_disponible * (pesos["lote"] / total),
            ancho_disponible * (pesos["pres"] / total),
            ancho_disponible * (pesos["exist"] / total),
            ancho_disponible * (pesos["costo"] / total),
            ancho_disponible * (pesos["cad"] / total),
        ]

        tabla = Table(data_tabla, colWidths=col_widths)

        
        # Estilo de tabla
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8B0000")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        # Obtener altura de la tabla
        wrap_height = tabla.wrapOn(p, width - 2*inch, height)[1]
        y_tabla = y_actual - wrap_height - 20
        
        # Si no cabe en la página, crear nueva página
        if y_tabla < (inch * 2.5):
            p.showPage()
            y_tabla = height - inch - wrap_height
        
        # Dibujar tabla
        tabla.drawOn(p, inch, y_tabla)
        
        # Pie de página
        p.setFont("Helvetica", 9)
        p.drawCentredString(width / 2.0, inch * 0.5, f"Documento generado por INVENTFARM - {nombre_usuario}")
        
        # Finalizar PDF
        p.showPage()
        p.save()
        
        # Retornar como descarga
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Inventario_{datetime.now().strftime("%d%m%Y")}.pdf"'
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error: {str(e)}', status=400)



@login_required
@require_http_methods(["POST"])
def actualizar_cpm(request):
    """Actualizar el CPM de un medicamento"""
    try:
        data = json.loads(request.body)
        medicamento_id = data.get('medicamento_id')
        cpm_valor = data.get('cpm')
        
        medicamento = Medicamento.objects.get(id=medicamento_id)
        
        # Actualizar o crear CPM
        cpm_obj, created = CPMMedicamento.objects.get_or_create(medicamento=medicamento)
        cpm_obj.valor = cpm_valor
        cpm_obj.save()
        
        return JsonResponse({'success': True})
    except Medicamento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Medicamento no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST"])
def eliminar_medicamento(request):
    """Eliminar un medicamento y todos sus lotes"""
    try:
        data = json.loads(request.body)
        medicamento_id = data.get('medicamento_id')
        
        medicamento = Medicamento.objects.get(id=medicamento_id)
        
        # Eliminar todos los lotes asociados
        Lote.objects.filter(medicamento=medicamento).delete()
        
        # Eliminar el medicamento
        medicamento.delete()
        
        return JsonResponse({'success': True})
    except Medicamento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Medicamento no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def exportar_inventario_general_excel(request):
    """Exportar inventario general a Excel"""
    try:
        import xlsxwriter
        
        # Obtener inventario general (agrupado por medicamento)
        inventario = Lote.objects.values(
            'medicamento__id',
            'medicamento__clave',
            'medicamento__descripcion',
        ).annotate(
            existencia_total=Sum('existencia'),
            cpm_medicamento=Coalesce(
                F('medicamento__cpm_medicamento__valor'),
                Value(0),
                output_field=IntegerField()
            )
        ).filter(
            existencia_total__gt=0
        ).order_by('medicamento__descripcion')
        
        # Crear archivo en memoria
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Inventario General')
        
        # Definir formatos
        header_format = workbook.add_format({
            'bg_color': '#8B0000',
            'font_color': 'white',
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_size': 11
        })
        
        title_format = workbook.add_format({
            'bg_color': '#8B0000',
            'font_color': 'white',
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 14
        })
        
        date_format = workbook.add_format({
            'italic': True,
            'align': 'left',
            'font_size': 10
        })
        
        data_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_size': 10
        })
        
        critico_format = workbook.add_format({
            'bg_color': '#FF0000',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'bold': True
        })
        
        bajo_format = workbook.add_format({
            'bg_color': '#FF4444',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        yellow_format = workbook.add_format({
            'bg_color': '#FFFF00',
            'font_color': 'black',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        green_format = workbook.add_format({
            'bg_color': '#00B050',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        # Establecer ancho de columnas
        worksheet.set_column('A:A', 18)
        worksheet.set_column('B:B', 35)
        worksheet.set_column('C:C', 18)
        worksheet.set_column('D:D', 15)
        worksheet.set_column('E:E', 15)
        
        # Agregar logo
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia', 'static', 'farmacia', 'img', 'logo.jpg')
        if os.path.exists(logo_path):
            try:
                worksheet.insert_image('A1', logo_path, {
                    'x_scale': 0.8,
                    'y_scale': 0.8,
                    'x_offset': 0,
                    'y_offset': 0
                })
            except Exception as e:
                print(f"Error insertando logo: {e}")
        
        # Título
        worksheet.merge_range('A3:E3', 'REPORTE DE INVENTARIO GENERAL', title_format)
        worksheet.set_row(2, 25)
        
        # Fecha
        worksheet.merge_range('A4:E4', f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_format)
        worksheet.set_row(3, 18)
        
        # Encabezados
        headers = ['Clave', 'Descripción', 'Existencia Total', 'CPM', 'Estado']
        for col, header in enumerate(headers):
            worksheet.write(5, col, header, header_format)
        worksheet.set_row(5, 20)
        
        # Datos
        row = 6
        for item in inventario:
            worksheet.write(row, 0, item['medicamento__clave'], data_format)
            worksheet.write(row, 1, item['medicamento__descripcion'], data_format)
            worksheet.write(row, 2, item['existencia_total'], data_format)
            worksheet.write(row, 3, item['cpm_medicamento'], data_format)
            
            # Estado según existencia
            existencia = item['existencia_total']
            if existencia <= 10:
                estado_text = 'Crítico'
                estado_format = critico_format
            elif existencia <= 50:
                estado_text = 'Bajo'
                estado_format = bajo_format
            elif existencia <= 100:
                estado_text = 'Medio'
                estado_format = yellow_format
            else:
                estado_text = 'Adecuado'
                estado_format = green_format
            
            worksheet.write(row, 4, estado_text, estado_format)
            worksheet.set_row(row, 18)
            row += 1
        
        # Pie de página
        worksheet.merge_range(row + 1, 0, row + 1, 4, 'Documento generado automáticamente por INVENTFARM', date_format)
        
        # Finalizar
        workbook.close()
        
        # Preparar respuesta
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Inventario_General_{datetime.now().strftime("%d%m%Y")}.xlsx"'
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error: {str(e)}', status=400)


@login_required
def exportar_inventario_general_pdf(request):
    """Exportar inventario general a PDF (descripción con wrap)"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT

        # Obtener inventario general
        inventario = Lote.objects.values(
            'medicamento__id',
            'medicamento__clave',
            'medicamento__descripcion',
        ).annotate(
            existencia_total=Sum('existencia'),
            cpm_medicamento=Coalesce(
                F('medicamento__cpm_medicamento__valor'),  # OJO: related_name en tu modelo es cpmmedicamento [file:5]
                Value(0),
                output_field=IntegerField()
            )
        ).filter(
            existencia_total__gt=0
        ).order_by('medicamento__descripcion')  # [file:6]

        # Crear PDF
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=landscape(letter))
        width, height = landscape(letter)

        # Logo
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia', 'static', 'farmacia', 'img', 'logo.jpg')
        logo_width = 7.0 * inch
        logo_height = 1.0 * inch
        x_logo = (width - logo_width) / 2
        y_logo = height - (0.75 * inch) - logo_height

        if os.path.exists(logo_path):
            try:
                p.drawImage(
                    logo_path,
                    x_logo, y_logo,
                    width=logo_width, height=logo_height,
                    preserveAspectRatio=True
                )
            except Exception as e:
                print(f"Error cargando logo: {e}")

        y_actual = y_logo - (0.25 * inch)

        # Título
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width / 2.0, y_actual, "REPORTE DE INVENTARIO GENERAL DE MEDICAMENTOS")
        y_actual -= 20

        # Fecha
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2.0, y_actual, f"Fecha del Reporte: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y_actual -= 20
        
                
        nombre_usuario = (request.user.get_full_name() or request.user.username).strip()
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2.0, y_actual, f"Generado por: {nombre_usuario}")
        y_actual -= 16
        

        # Línea
        p.line(inch, y_actual, width - inch, y_actual)
        y_actual -= 20

        # ====== FIX: Paragraph para wrap en descripción ======
        styles = getSampleStyleSheet()
        desc_style = ParagraphStyle(
            "desc",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            leading=8,
            alignment=TA_LEFT,
        )

        # Tabla
        data_tabla = [['Clave', 'Descripción', 'Existencia', 'CPM', 'Estado']]

        for item in inventario:
            existencia = item['existencia_total']
            if existencia <= 10:
                estado = 'Crítico'
            elif existencia <= 50:
                estado = 'Bajo'
            elif existencia <= 100:
                estado = 'Medio'
            else:
                estado = 'Adecuado'

            data_tabla.append([
                item['medicamento__clave'],
                Paragraph(item['medicamento__descripcion'] or "", desc_style),  # ✅ sin recorte
                str(existencia),
                str(item['cpm_medicamento']),
                estado
            ])

        margen_x = inch
        ancho_disponible = width - (2 * margen_x)

        # Reparto proporcional (ajusta estos “pesos” a tu gusto)
        pesos = {
            "clave": 1.2,
            "descripcion": 6.0,   # la más grande
            "existencia": 1.1,
            "cpm": 0.9,
            "estado": 1.1,
        }

        total_pesos = sum(pesos.values())
        col_widths = [
            ancho_disponible * (pesos["clave"] / total_pesos),
            ancho_disponible * (pesos["descripcion"] / total_pesos),
            ancho_disponible * (pesos["existencia"] / total_pesos),
            ancho_disponible * (pesos["cpm"] / total_pesos),
            ancho_disponible * (pesos["estado"] / total_pesos),
        ]

        tabla = Table(data_tabla, colWidths=col_widths)

        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

            # Cuerpo
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            # ✅ Para que el texto envuelto se vea bien
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('LEFTPADDING', (1, 1), (1, -1), 4),
            ('RIGHTPADDING', (1, 1), (1, -1), 4),
        ]))

        # Dibujar tabla
        wrap_height = tabla.wrapOn(p, width - 2*inch, height)[1]
        y_tabla = y_actual - wrap_height - 20

        if y_tabla < (inch * 2.5):
            p.showPage()
            y_tabla = height - inch - wrap_height

        tabla.drawOn(p, inch, y_tabla)

        # Pie
        p.setFont("Helvetica", 9)
        p.drawCentredString(width / 2.0, inch * 0.5, "Documento generado por INVENTFARM")

        p.showPage()
        p.save()

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Inventario_General_{datetime.now().strftime("%d%m%Y")}.pdf"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error: {str(e)}', status=400)

    
# ===== MÓDULO DE REPORTES =====

@login_required
@permission_required('farmacia.view_reportes', raise_exception=True)
def reportes_farmacia(request):
    """Vista principal del módulo de reportes"""
    return render(request, 'reportes.html', {
        'user': request.user
    })


@login_required
def api_reportes_kpis(request):
    try:
        fecha_fin = timezone.now().date()
        fecha_inicio = fecha_fin - timedelta(days=90)

        if request.GET.get('fecha_inicio'):
            fecha_inicio = datetime.strptime(request.GET.get('fecha_inicio'), '%Y-%m-%d').date()
        if request.GET.get('fecha_fin'):
            fecha_fin = datetime.strptime(request.GET.get('fecha_fin'), '%Y-%m-%d').date()

        total_salidas = RecetaMedicamento.objects.filter(
            receta__fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).count()

        total_medicamentos = RecetaMedicamento.objects.filter(
            receta__fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).aggregate(total=Sum('cantidad_surtida'))['total'] or 0

        total_pacientes = Receta.objects.filter(
            fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).values('paciente').distinct().count()

        valor_total = RecetaMedicamento.objects.filter(
            receta__fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).aggregate(total=Sum('precio_total'))['total'] or 0

        return JsonResponse({
            'success': True,
            'kpis': {
                'total_salidas': total_salidas,
                'total_medicamentos': total_medicamentos,
                'total_pacientes': total_pacientes,
                'valor_total': float(valor_total),
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def api_reportes_salidas(request):
    """API para obtener datos de salidas para reportes"""
    try:
        fecha_fin = timezone.now().date()
        fecha_inicio = fecha_fin - timedelta(days=90)
        
        if request.GET.get('fecha_inicio'):
            fecha_inicio = datetime.strptime(request.GET.get('fecha_inicio'), '%Y-%m-%d').date()
        if request.GET.get('fecha_fin'):
            fecha_fin = datetime.strptime(request.GET.get('fecha_fin'), '%Y-%m-%d').date()
        
        salidas = RecetaMedicamento.objects.filter(
            receta__fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).select_related(
            'receta__paciente',
            'receta__surtido_por',
            'medicamento',
            'lote'
        ).order_by('-receta__fecha_surtido')[:100]
        
        datos_salidas = []
        for item in salidas:
            if isinstance(item.receta.fecha_surtido, datetime):
                fecha_str = item.receta.fecha_surtido.strftime('%Y-%m-%d')
                hora_str = item.receta.fecha_surtido.strftime('%H:%M')
            else:
                fecha_str = item.receta.fecha_surtido.strftime('%Y-%m-%d')
                hora_str = '--:--'
            
            if item.receta.id_folio.startswith('COL-'):
                tipo = 'Colectivo - Paciente'
                tipo_badge = 'badge-colectivo'
            elif item.receta.id_folio.startswith('STK-'):
                tipo = 'Colectivo - Stock'
                tipo_badge = 'badge-stock'
            else:
                tipo = 'Receta'
                tipo_badge = 'badge-receta'
            
            paciente_nombre = 'Stock/Servicio'
            if item.receta.paciente:
                paciente_nombre = item.receta.paciente.nombre_completo
            
            # ✅ MEJORAR: Obtener nombre del responsable
            responsable_nombre = 'N/A'
            if item.receta.surtido_por:
                # Intentar usar first_name y last_name
                nombre_completo = f"{item.receta.surtido_por.first_name} {item.receta.surtido_por.last_name}".strip()
                # Si está vacío, usar username
                responsable_nombre = nombre_completo if nombre_completo else item.receta.surtido_por.username
            
            datos_salidas.append({
                'id': item.receta.id,
                'folio': item.receta.id_folio,
                'fecha': fecha_str,
                'hora': hora_str,
                'medicamento': item.medicamento.descripcion,
                'cantidad': item.cantidad_surtida,
                'paciente': paciente_nombre,
                'responsable': responsable_nombre,  # ✅ Ahora siempre tendrá un valor
                'valor': float(item.precio_total or 0),
                'precio_unitario': float(item.precio_unitario or 0),
                'tipo': tipo,
                'tipo_badge': tipo_badge
            })
        
        return JsonResponse({
            'success': True,
            'data': datos_salidas,
            'total': len(datos_salidas)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




@login_required
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def api_reportes_medicamentos_top(request):
    """API para obtener medicamentos más dispensados"""
    try:
        # Obtener rango de fechas
        fecha_fin = timezone.now().date()
        fecha_inicio = fecha_fin - timedelta(days=90)
        
        if request.GET.get('fecha_inicio'):
            fecha_inicio = datetime.strptime(request.GET.get('fecha_inicio'), '%Y-%m-%d').date()
        if request.GET.get('fecha_fin'):
            fecha_fin = datetime.strptime(request.GET.get('fecha_fin'), '%Y-%m-%d').date()
        
        # Agrupar y contar - CORREGIDO
        medicamentos_top = RecetaMedicamento.objects.filter(
            receta__fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).values(
            'medicamento__id',
            'medicamento__descripcion',
            'medicamento__clave'
        ).annotate(
            total_dispensado=Sum('cantidad_surtida')
        ).order_by('-total_dispensado')[:10]
        
        datos = []
        for med in medicamentos_top:
            datos.append({
                'medicamento': med['medicamento__descripcion'],
                'clave': med['medicamento__clave'],
                'cantidad': med['total_dispensado']
            })
        
        return JsonResponse({
            'success': True,
            'data': datos
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def api_reportes_pacientes_frecuentes(request):
    """API para obtener pacientes más atendidos"""
    try:
        fecha_fin = timezone.now().date()
        fecha_inicio = fecha_fin - timedelta(days=90)
        
        if request.GET.get('fecha_inicio'):
            fecha_inicio = datetime.strptime(request.GET.get('fecha_inicio'), '%Y-%m-%d').date()
        if request.GET.get('fecha_fin'):
            fecha_fin = datetime.strptime(request.GET.get('fecha_fin'), '%Y-%m-%d').date()
        
        # Agrupar por paciente - CORREGIDO
        pacientes_top = Receta.objects.filter(
            fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).values(
            'paciente__id',
            'paciente__nombre_completo'
        ).annotate(
            total_visitas=Count('id'),
            total_medicamentos=Sum('recetamedicamento__cantidad_surtida'),
            ultima_visita=Max('fecha_surtido')
        ).order_by('-total_visitas')[:10]
        
        datos = []
        for pac in pacientes_top:
            # Manejar fecha que puede ser DateField o DateTimeField
            if pac['ultima_visita']:
                if isinstance(pac['ultima_visita'], datetime):
                    ultima_visita_str = pac['ultima_visita'].strftime('%Y-%m-%d')
                else:
                    ultima_visita_str = pac['ultima_visita'].strftime('%Y-%m-%d')
            else:
                ultima_visita_str = 'N/A'
            
            datos.append({
                'paciente': pac['paciente__nombre_completo'],
                'visitas': pac['total_visitas'],
                'medicamentos': pac['total_medicamentos'] or 0,
                'ultima_visita': ultima_visita_str,
                'gasto_total': 0
            })
        
        return JsonResponse({
            'success': True,
            'data': datos
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def api_reportes_tendencias(request):
    """API para obtener tendencias mensuales"""
    try:
        # Últimos 12 meses
        fecha_fin = timezone.now().date()
        fecha_inicio = fecha_fin - timedelta(days=365)
        
        # Agrupar por mes - CORREGIDO
        salidas_por_mes = RecetaMedicamento.objects.filter(
            receta__fecha_surtido__range=[fecha_inicio, fecha_fin]
        ).annotate(
            mes=TruncMonth('receta__fecha_surtido')
        ).values('mes').annotate(
            total=Count('id')
        ).order_by('mes')
        
        meses = []
        totales = []
        
        for item in salidas_por_mes:
            meses.append(item['mes'].strftime('%b %Y'))
            totales.append(item['total'])
        
        return JsonResponse({
            'success': True,
            'meses': meses,
            'totales': totales
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        

# ===== COLECTIVOS - FARMACIA (AGREGAR AL FINAL) =====


@never_cache
@login_required(login_url='login')
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def lista_colectivos_farmacia(request):
    """Vista de lista de colectivos para farmacia"""
    # Obtener todos los colectivos
    colectivos = Colectivo.objects.select_related(
        'paciente', 
        'enfermero_solicitante',
        'farmaceutico_asignado'
    ).order_by('-fecha_solicitud')
    
    # Filtros
    estado_filtro = request.GET.get('estado', '')
    busqueda = request.GET.get('q', '')
    
    if estado_filtro:
        colectivos = colectivos.filter(estado=estado_filtro)
    
    if busqueda:
        colectivos = colectivos.filter(
            Q(folio__icontains=busqueda) |
            Q(paciente__nombre__icontains=busqueda) |
            Q(paciente__apellido_paterno__icontains=busqueda) |
            Q(paciente__apellido_materno__icontains=busqueda) |
            Q(numero_cama__icontains=busqueda) |
            Q(servicio__icontains=busqueda) |
            Q(enfermero_solicitante__username__icontains=busqueda)
        )
    
    # Estadísticas - SOLUCIÓN CON RANGO DE FECHAS
    now_local = timezone.localtime(timezone.now())
    hoy_inicio = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    hoy_fin = now_local.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    stats = {
        'total': Colectivo.objects.count(),
        'pendientes': Colectivo.objects.filter(estado='PENDIENTE').count(),
        'en_revision': Colectivo.objects.filter(estado='EN_REVISION').count(),
        'respondidos': Colectivo.objects.filter(estado='RESPONDIDO').count(),
        'completados_hoy': Colectivo.objects.filter(
            estado='COMPLETADO',
            fecha_completado__gte=hoy_inicio,  # Mayor o igual a 00:00:00 de hoy
            fecha_completado__lte=hoy_fin      # Menor o igual a 23:59:59 de hoy
        ).count(),
    }
    
    return render(request, 'lista_colectivos_farmacia.html', {
        'colectivos': colectivos,
        'stats': stats,
        'estado_filtro': estado_filtro,
        'busqueda': busqueda,
    })
    

@never_cache
@login_required(login_url='login')
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia')
def detalle_colectivo_farmacia(request, colectivo_id):
    """Vista de detalle de un colectivo para farmacia"""
    colectivo = get_object_or_404(
        Colectivo.objects.select_related('paciente', 'enfermero_solicitante'),
        id=colectivo_id
    )
    
    if colectivo.estado == 'PENDIENTE':
        colectivo.estado = 'EN_REVISION'
        colectivo.farmaceutico_asignado = request.user
        colectivo.save()
    
    medicamentos = colectivo.medicamentos.select_related('medicamento').all()
    
    medicamentos_con_stock = []
    for item in medicamentos:
        stock_total = Lote.objects.filter(
            medicamento=item.medicamento,
            existencia__gt=0
        ).aggregate(total=Sum('existencia'))['total'] or 0
        
        medicamentos_con_stock.append({
            'item': item,
            'stock_disponible': stock_total,
            'suficiente': stock_total >= item.cantidad_solicitada
        })
    
    return render(request, 'detalle_colectivo_farmacia.html', {
        'colectivo': colectivo,
        'medicamentos_con_stock': medicamentos_con_stock,
        'user': request.user
    })


@never_cache
@login_required(login_url='login')
@require_http_methods(['POST'])
@permission_required('enfermeria.respond_colectivo', raise_exception=True)
def responder_colectivo(request, colectivo_id):
    """Farmacia responde al colectivo indicando disponibilidad"""
    colectivo = get_object_or_404(Colectivo, id=colectivo_id)
    
    if colectivo.estado not in ['PENDIENTE', 'EN_REVISION']:
        messages.error(request, 'Este colectivo ya fue respondido o completado')
        return redirect('detalle_colectivo_farmacia', colectivo_id=colectivo.id)
    
    try:
        colectivo.respuesta_farmacia = request.POST.get('respuesta_farmacia', '')
        
        for medicamento in colectivo.medicamentos.all():
            disponible = request.POST.get(f'disponible_{medicamento.id}') == 'on'
            comentario = request.POST.get(f'comentario_{medicamento.id}', '')
            
            medicamento.disponible = disponible
            medicamento.comentario_farmacia = comentario
            medicamento.save()
        
        colectivo.estado = 'RESPONDIDO'
        colectivo.fecha_respuesta_farmacia = timezone.now()
        colectivo.farmaceutico_asignado = request.user
        colectivo.save()
        
        messages.success(request, f'Respuesta enviada a enfermería para colectivo {colectivo.folio}')
        return redirect('lista_colectivos_farmacia')
        
    except Exception as e:
        messages.error(request, f'Error al responder colectivo: {str(e)}')
        return redirect('detalle_colectivo_farmacia', colectivo_id=colectivo.id)



@never_cache
@login_required(login_url='login')
@require_http_methods(['POST'])
@permission_required('enfermeria.complete_colectivo', raise_exception=True)
def completar_colectivo(request, colectivo_id):
    """Marca el colectivo como completado y descuenta del inventario"""
    colectivo = get_object_or_404(Colectivo, id=colectivo_id)
    
    if colectivo.estado != 'EN_REVISION':
        messages.error(request, 'Solo se pueden completar colectivos en revisión')
        return redirect('detalle_colectivo_farmacia', colectivo_id=colectivo.id)
    
    try:
        with transaction.atomic():
            # PRIMERO: Validar stock
            for medicamento in colectivo.medicamentos.all():
                cantidad_surtida = int(request.POST.get(f'cantidad_surtida_{medicamento.id}', 0))
                
                stock_total = Lote.objects.filter(
                    medicamento=medicamento.medicamento,
                    existencia__gt=0
                ).aggregate(total=Sum('existencia'))['total'] or 0
                
                if cantidad_surtida > stock_total:
                    messages.error(
                        request, 
                        f'Stock insuficiente para {medicamento.medicamento.descripcion}. '
                        f'Disponible: {stock_total}, Solicitado: {cantidad_surtida}'
                    )
                    return redirect('detalle_colectivo_farmacia', colectivo_id=colectivo.id)
            
            # SEGUNDO: Actualizar cantidades surtidas
            for medicamento in colectivo.medicamentos.all():
                cantidad_surtida = int(request.POST.get(f'cantidad_surtida_{medicamento.id}', 0))
                medicamento.cantidad_surtida = cantidad_surtida
                medicamento.save()
            
            # DEBUG
            print("="*60)
            print("🔍 COMPLETANDO COLECTIVO")
            print(f"   Folio: {colectivo.folio}")
            print(f"   Tipo: {colectivo.tipo_colectivo}")
            print(f"   Paciente: {colectivo.paciente}")
            
            # TERCERO: ✅ Crear Receta para AMBOS tipos
            receta = None

            if colectivo.tipo_colectivo == 'PACIENTE':
                if not colectivo.paciente:
                    raise ValueError('Colectivo de tipo PACIENTE debe tener un paciente asignado')
                
                print("   📝 Creando receta para paciente...")
                
                receta = Receta.objects.create(
                    id_folio=f"{colectivo.folio}",
                    paciente=colectivo.paciente,
                    fecha_emision=colectivo.fecha_solicitud.date() if hasattr(colectivo.fecha_solicitud, 'date') else colectivo.fecha_solicitud,
                    fecha_surtido=timezone.now().date(),
                    estado='completa',
                    origen='hospitalizacion_adultos',
                    surtido_por=request.user
                )
                print(f"   ✅ Receta creada para paciente: {receta.id_folio} (ID: {receta.id})")

            elif colectivo.tipo_colectivo == 'STOCK':
                print("   📝 Creando receta para colectivo de STOCK...")
                
                # ✅ NUEVO: Crear un paciente genérico para stock SI NO EXISTE
                paciente_stock, created = Paciente.objects.get_or_create(
                    curp='STCK000000HDFXXX00',
                    defaults={
                        'nombre_completo': f'STOCK - {colectivo.servicio}',
                        'fecha_nacimiento': date(1900, 1, 1)
                    }
                )
                
                receta = Receta.objects.create(
                    id_folio=f"{colectivo.folio}",  # Ya tiene prefijo STK-
                    paciente=paciente_stock,  # ✅ Usa paciente genérico
                    fecha_emision=colectivo.fecha_solicitud.date() if hasattr(colectivo.fecha_solicitud, 'date') else colectivo.fecha_solicitud,
                    fecha_surtido=timezone.now().date(),
                    estado='completa',
                    origen='hospitalizacion_adultos',
                    surtido_por=request.user
                )
                print(f"   ✅ Receta creada para STOCK: {receta.id_folio} (ID: {receta.id})")
            else:
                print("   ℹ️  Colectivo de tipo STOCK - No se genera receta")
            
            # CUARTO: Descontar del inventario Y CALCULAR PRECIOS
            medicamentos_registrados = 0
            for medicamento in colectivo.medicamentos.all():
                cantidad_restante = medicamento.cantidad_surtida
                
                print(f"\n📦 Procesando: {medicamento.medicamento.descripcion}")
                print(f"   Cantidad a surtir: {medicamento.cantidad_surtida}")
                
                lotes = Lote.objects.filter(
                    medicamento=medicamento.medicamento,
                    existencia__gt=0
                ).order_by('fecha_caducidad')
                
                lote_usado = None
                precio_acumulado = Decimal('0.00')  # ✅ NUEVO
                
                for lote in lotes:
                    if cantidad_restante <= 0:
                        break
                    
                    if not lote_usado:
                        lote_usado = lote
                    
                    # ✅ CALCULAR PRECIO SEGÚN CANTIDAD TOMADA DE ESTE LOTE
                    if lote.existencia >= cantidad_restante:
                        precio_acumulado += (lote.costo_unitario * cantidad_restante)  # ✅ NUEVO
                        lote.existencia -= cantidad_restante
                        lote.save()
                        cantidad_restante = 0
                    else:
                        precio_acumulado += (lote.costo_unitario * lote.existencia)  # ✅ NUEVO
                        cantidad_restante -= lote.existencia
                        lote.existencia = 0
                        lote.save()
                
                # ✅ CALCULAR PRECIO UNITARIO PROMEDIO
                precio_unitario_promedio = precio_acumulado / medicamento.cantidad_surtida if medicamento.cantidad_surtida > 0 else Decimal('0.00')
                
                # ✅ Crear RecetaMedicamento SOLO si se creó una receta
                if receta:
                    try:
                        receta_med = RecetaMedicamento.objects.create(
                            receta=receta,
                            medicamento=medicamento.medicamento,
                            lote=lote_usado,
                            cantidad_solicitada=medicamento.cantidad_solicitada,
                            cantidad_surtida=medicamento.cantidad_surtida,
                            precio_unitario=precio_unitario_promedio,  # ✅ NUEVO
                            precio_total=precio_acumulado  # ✅ NUEVO
                        )
                        medicamentos_registrados += 1
                        print(f"   ✅ RecetaMedicamento creado (ID: {receta_med.id}) - Precio: ${precio_acumulado}")
                    except Exception as e:
                        print(f"   ❌ ERROR al crear RecetaMedicamento: {str(e)}")
                        raise
                else:
                    print(f"   ⏭️  No se registra en RecetaMedicamento (colectivo de stock)")

            
            if receta:
                print(f"\n✅ TOTAL MEDICAMENTOS REGISTRADOS EN RECETA: {medicamentos_registrados}")
            
            # QUINTO: Cambiar estado a COMPLETADO
            colectivo.estado = 'COMPLETADO'
            colectivo.fecha_completado = timezone.now()
            colectivo.farmaceutico_asignado = request.user
            colectivo.save()
            
            print(f"✅ Colectivo completado: {colectivo.folio}")
            print("="*60)
        
        messages.success(
            request, 
            f'Colectivo {colectivo.folio} completado exitosamente.'
        )
        
        return redirect('lista_colectivos_farmacia')
        
    except Exception as e:
        import traceback
        print(f"\n❌ ERROR GENERAL: {str(e)}")
        traceback.print_exc()
        messages.error(request, f'Error al completar colectivo: {str(e)}')
        return redirect('detalle_colectivo_farmacia', colectivo_id=colectivo.id)




@login_required(login_url='login')
@group_required('Administrador', 'Farmacéutico', 'Jefe de Farmacia', 'Enfermero', 'Jefe de Enfermería')
def generar_pdf_colectivo(request, colectivo_id):
    """Genera PDF con la información del colectivo completado"""
    colectivo = get_object_or_404(
        Colectivo.objects.select_related('paciente', 'enfermero_solicitante', 'farmaceutico_asignado'),
        id=colectivo_id
    )
    
    if colectivo.estado != 'COMPLETADO':
        messages.error(request, 'Solo se puede generar PDF de colectivos completados')
        return redirect('detalle_colectivo_farmacia', colectivo_id=colectivo.id)
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
        from io import BytesIO
        import os
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Logo
        logo_path = os.path.join(settings.BASE_DIR, 'farmacia', 'static', 'farmacia', 'img', 'logo.jpg')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=3.5*inch, height=0.5*inch)
                elements.append(logo)
                elements.append(Spacer(1, 0.2*inch))
            except:
                pass
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#750000'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        title = Paragraph(f"COLECTIVO DE MEDICAMENTOS<br/>{colectivo.folio}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.15*inch))
        
        # DETECTAR TIPO DE COLECTIVO
        es_colectivo_stock = colectivo.paciente is None
        
        # Información según el tipo de colectivo
        if es_colectivo_stock:
            # COLECTIVO DE STOCK - Sin información de paciente
            info_data = [
                ['INFO. DEL COLECTIVO', ''],
                ['Tipo:', 'Colectivo para Stock'],
                ['Servicio:', colectivo.servicio or 'N/A'],
                ['Número de Cama:', colectivo.numero_cama or 'N/A'],
                ['', ''],
                ['FECHAS Y RESPONSABLES', ''],
                ['Fecha de Solicitud:', timezone.localtime(colectivo.fecha_solicitud).strftime('%d/%m/%Y %H:%M') if colectivo.fecha_solicitud else 'N/A'],
                ['Fecha de Surtido:', timezone.localtime(colectivo.fecha_completado).strftime('%d/%m/%Y %H:%M') if colectivo.fecha_completado else 'N/A'],
                ['Enfermero(a) Solicitante:', colectivo.enfermero_solicitante.get_full_name() or colectivo.enfermero_solicitante.username if colectivo.enfermero_solicitante else 'N/A'],
                ['Farmacéutico(a) Asignado:', colectivo.farmaceutico_asignado.get_full_name() or colectivo.farmaceutico_asignado.username if colectivo.farmaceutico_asignado else 'N/A'],
            ]
        else:
            # COLECTIVO DE PACIENTE - Con información completa del paciente
            info_data = [
                ['INFORMACIÓN DEL PACIENTE', ''],
                ['Nombre:', colectivo.paciente.nombre_completo],
                ['CURP:', colectivo.paciente.curp if colectivo.paciente.curp else 'N/A'],
                ['Fecha de Nacimiento:', colectivo.paciente.fecha_nacimiento.strftime('%d/%m/%Y') if colectivo.paciente.fecha_nacimiento else 'N/A'],
                ['Número de Cama:', colectivo.numero_cama or 'N/A'],
                ['Servicio:', colectivo.servicio or 'N/A'],
                ['', ''],
                ['INFO. DEL COLECTIVO', ''],
                ['Fecha de Solicitud:', timezone.localtime(colectivo.fecha_solicitud).strftime('%d/%m/%Y %H:%M') if colectivo.fecha_solicitud else 'N/A'],
                ['Fecha de Surtido:', timezone.localtime(colectivo.fecha_completado).strftime('%d/%m/%Y %H:%M') if colectivo.fecha_completado else 'N/A'],
                ['Enfermero(a):', colectivo.enfermero_solicitante.get_full_name() or colectivo.enfermero_solicitante.username if colectivo.enfermero_solicitante else 'N/A'],
                ['Farmacéutico(a):', colectivo.farmaceutico_asignado.get_full_name() or colectivo.farmaceutico_asignado.username if colectivo.farmaceutico_asignado else 'N/A'],
            ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#750000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, 5 if es_colectivo_stock else 7), (-1, 5 if es_colectivo_stock else 7), colors.HexColor('#750000')),
            ('TEXTCOLOR', (0, 5 if es_colectivo_stock else 7), (-1, 5 if es_colectivo_stock else 7), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 5 if es_colectivo_stock else 7), (-1, 5 if es_colectivo_stock else 7), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # ESTILO PARA TEXTO EN CELDAS (con word wrap)
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT
        )
        
        # Estilo centrado para números
        center_style = ParagraphStyle(
            'CenterStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=TA_CENTER
        )
        
        # Tabla de medicamentos - CON PARAGRAPH PARA WORD WRAP
        medicamentos_data = [['#', 'CLAVE', 'DESCRIPCIÓN', 'SOLICITADO', 'SURTIDO']]
        
        for idx, item in enumerate(colectivo.medicamentos.all(), 1):
            # Usar Paragraph para todos los campos
            numero_paragraph = Paragraph(str(idx), center_style)
            clave_paragraph = Paragraph(item.medicamento.clave if item.medicamento else 'N/A', cell_style)
            descripcion_paragraph = Paragraph(item.medicamento.descripcion if item.medicamento else 'N/A', cell_style)
            solicitado_paragraph = Paragraph(str(item.cantidad_solicitada), center_style)
            surtido_paragraph = Paragraph(str(item.cantidad_surtida), center_style)
            
            medicamentos_data.append([
                numero_paragraph,
                clave_paragraph,
                descripcion_paragraph,
                solicitado_paragraph,
                surtido_paragraph
            ])
        
        # Tabla con anchos ajustados
        medicamentos_table = Table(
            medicamentos_data, 
            colWidths=[0.3*inch, 1.2*inch, 3.8*inch, 0.9*inch, 0.9*inch],
            repeatRows=1
        )
        medicamentos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#750000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(medicamentos_table)
        
        # Observaciones
        if colectivo.observaciones_enfermeria:
            elements.append(Spacer(1, 0.15*inch))
            obs_style = ParagraphStyle('Observaciones', parent=styles['Normal'], fontSize=8)
            elements.append(Paragraph(f"<b>Observaciones de Enfermería:</b> {colectivo.observaciones_enfermeria}", obs_style))
        
        # Construir PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        
        # Enviar respuesta
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Colectivo_{colectivo.folio}.pdf"'
        response.write(pdf)
        
        return response
        
    except Exception as e:
        import traceback
        print(f"Error al generar PDF: {str(e)}")
        print(traceback.format_exc())
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('lista_colectivos_farmacia')


"""
===== ADMINISTRACIÓN DE USUARIOS Y GRUPOS =====
Vistas para gestionar usuarios y grupos sin usar el admin de Django.
Solo accesible para usuarios del grupo 'Administrador'.
"""


@login_required
@group_required('Administrador')
def admin_usuarios(request):
    """
    Vista principal para listar todos los usuarios del sistema.
    Muestra estadísticas y permite filtrar por grupo y estado.
    """
    usuarios = User.objects.all().prefetch_related('groups').order_by('-date_joined')
    grupos = Group.objects.all().order_by('name')
    
    # Calcular estadísticas
    usuarios_activos = usuarios.filter(is_active=True, is_superuser=False).count()
    total_grupos = grupos.count()
    total_admins = usuarios.filter(Q(is_superuser=True) | Q(groups__name='Administrador')).distinct().count()
    
    context = {
        'usuarios': usuarios,
        'grupos': grupos,
        'usuarios_activos': usuarios_activos,
        'total_grupos': total_grupos,
        'total_admins': total_admins,
    }
    
    return render(request, 'admin_usuarios.html', context)


@login_required
@group_required('Administrador')
def admin_usuario_detalle(request, user_id):
    """
    Vista para ver y editar los detalles de un usuario.
    Permite cambiar datos básicos, grupos y contraseña.
    """
    usuario = get_object_or_404(User, pk=user_id)
    grupos = Group.objects.all().order_by('name')
    
    if request.method == 'POST':
        try:
            # 1. ACTUALIZAR DATOS BÁSICOS
            usuario.username = request.POST.get('username', usuario.username)
            usuario.email = request.POST.get('email', usuario.email)
            usuario.first_name = request.POST.get('first_name', usuario.first_name)
            usuario.last_name = request.POST.get('last_name', usuario.last_name)
            
            # 2. ACTUALIZAR ESTADO ACTIVO
            # Nota: los checkboxes no envían nada si están desmarcados
            usuario.is_active = request.POST.get('is_active') == 'on'
            
            # 3. ACTUALIZAR CONTRASEÑA (solo si se proporcionó)
            new_password = request.POST.get('new_password')
            if new_password and new_password.strip():
                usuario.set_password(new_password)
            
            # 4. GUARDAR USUARIO
            usuario.save()
            
            # 5. ACTUALIZAR GRUPOS
            grupos_seleccionados = request.POST.getlist('groups')
            usuario.groups.clear()  # Limpiar grupos actuales
            
            if grupos_seleccionados:
                for grupo_id in grupos_seleccionados:
                    try:
                        grupo = Group.objects.get(pk=grupo_id)
                        usuario.groups.add(grupo)
                    except Group.DoesNotExist:
                        continue
            
            messages.success(request, f'Usuario {usuario.username} actualizado exitosamente')
            return redirect('admin_usuarios')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')
    
    # Obtener permisos del usuario (de todos sus grupos)
    permisos_usuario = usuario.get_all_permissions()
    
    context = {
        'usuario': usuario,
        'grupos': grupos,
        'permisos_usuario': sorted(permisos_usuario),
    }
    
    return render(request, 'admin_usuario_detalle.html', context)



@login_required
@group_required('Administrador')
def admin_crear_usuario(request):
    """
    Vista para crear un nuevo usuario.
    Valida que el username sea único y crea el usuario con los grupos seleccionados.
    """
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            email = request.POST.get('email', '')
            first_name = request.POST.get('first_name', '')
            last_name = request.POST.get('last_name', '')
            password = request.POST.get('password')
            password2 = request.POST.get('password2')
            
            # Validaciones
            if not username or not password:
                messages.error(request, 'El nombre de usuario y contraseña son obligatorios')
                return redirect('admin_usuarios')
            
            if password != password2:
                messages.error(request, 'Las contraseñas no coinciden')
                return redirect('admin_usuarios')
            
            if User.objects.filter(username=username).exists():
                messages.error(request, f'El usuario {username} ya existe')
                return redirect('admin_usuarios')
            
            # Crear usuario
            usuario = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Asignar grupos
            grupos_seleccionados = request.POST.getlist('groups')
            for grupo_id in grupos_seleccionados:
                grupo = Group.objects.get(pk=grupo_id)
                usuario.groups.add(grupo)
            
            messages.success(request, f'Usuario {username} creado exitosamente')
            return redirect('admin_usuarios')
            
        except Exception as e:
            messages.error(request, f'Error al crear usuario: {str(e)}')
            return redirect('admin_usuarios')
    
    return redirect('admin_usuarios')


@login_required
@group_required('Administrador')
@require_http_methods(['POST'])
def admin_eliminar_usuario(request, user_id):
    """
    Vista para eliminar un usuario del sistema.
    No permite eliminar superusuarios ni el propio usuario.
    Retorna JSON para AJAX.
    """
    try:
        usuario = get_object_or_404(User, pk=user_id)
        
        # Validaciones de seguridad
        if usuario.is_superuser:
            return JsonResponse({
                'success': False,
                'error': 'No se puede eliminar un superusuario'
            }, status=403)
        
        if usuario.id == request.user.id:
            return JsonResponse({
                'success': False,
                'error': 'No puedes eliminar tu propio usuario'
            }, status=403)
        
        username = usuario.username
        usuario.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Usuario {username} eliminado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@group_required('Administrador')
def admin_grupos(request):
    """
    Vista para listar todos los grupos del sistema.
    Muestra cuántos usuarios tiene cada grupo y sus permisos.
    """
    grupos = Group.objects.annotate(
        num_usuarios=Count('user')
    ).prefetch_related('permissions').order_by('name')
    
    context = {
        'grupos': grupos,
    }
    
    return render(request, 'admin_grupos.html', context)


@login_required
@group_required('Administrador')
def admin_grupo_detalle(request, grupo_id):
    """
    Vista para ver y editar los detalles de un grupo.
    Permite cambiar nombre, permisos y ver usuarios asignados.
    """
    grupo = get_object_or_404(Group, pk=grupo_id)
    todos_permisos = Permission.objects.filter(
        content_type__app_label__in=['farmacia', 'enfermeria', 'auth']
    ).select_related('content_type').order_by('content_type__app_label', 'codename')
    
    if request.method == 'POST':
        try:
            # Actualizar nombre del grupo
            nuevo_nombre = request.POST.get('name', grupo.name)
            if nuevo_nombre != grupo.name:
                grupo.name = nuevo_nombre
                grupo.save()
            
            # Actualizar permisos
            # Importante: obtener la lista de permisos seleccionados
            permisos_seleccionados = request.POST.getlist('permissions')
            
            # Limpiar permisos actuales
            grupo.permissions.clear()
            
            # Agregar nuevos permisos
            if permisos_seleccionados:  # Solo si hay permisos seleccionados
                for permiso_id in permisos_seleccionados:
                    try:
                        permiso = Permission.objects.get(pk=permiso_id)
                        grupo.permissions.add(permiso)
                    except Permission.DoesNotExist:
                        continue
            
            messages.success(request, f'Grupo {grupo.name} actualizado exitosamente')
            return redirect('admin_grupos')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar grupo: {str(e)}')
    
    # Usuarios en este grupo
    usuarios_grupo = User.objects.filter(groups=grupo).order_by('username')
    
    # Permisos actuales del grupo
    permisos_grupo = grupo.permissions.values_list('id', flat=True)
    
    context = {
        'grupo': grupo,
        'todos_permisos': todos_permisos,
        'permisos_grupo': list(permisos_grupo),
        'usuarios_grupo': usuarios_grupo,
    }
    
    return render(request, 'admin_grupo_detalle.html', context)


@login_required
@group_required('Administrador')
def admin_crear_grupo(request):
    """
    Vista para crear un nuevo grupo.
    Valida que el nombre sea único.
    """
    if request.method == 'POST':
        try:
            nombre = request.POST.get('name')
            
            if not nombre:
                messages.error(request, 'El nombre del grupo es obligatorio')
                return redirect('admin_grupos')
            
            if Group.objects.filter(name=nombre).exists():
                messages.error(request, f'El grupo {nombre} ya existe')
                return redirect('admin_grupos')
            
            # Crear grupo
            grupo = Group.objects.create(name=nombre)
            
            # Asignar permisos si se seleccionaron
            permisos_seleccionados = request.POST.getlist('permissions')
            for permiso_id in permisos_seleccionados:
                permiso = Permission.objects.get(pk=permiso_id)
                grupo.permissions.add(permiso)
            
            messages.success(request, f'Grupo {nombre} creado exitosamente')
            return redirect('admin_grupos')
            
        except Exception as e:
            messages.error(request, f'Error al crear grupo: {str(e)}')
    
    return redirect('admin_grupos')


@login_required
@group_required('Administrador')
@require_http_methods(['POST'])
def admin_eliminar_grupo(request, grupo_id):
    """
    Vista para eliminar un grupo.
    Retorna JSON para AJAX.
    """
    try:
        grupo = get_object_or_404(Group, pk=grupo_id)
        nombre = grupo.name
        
        # Verificar que no tenga usuarios asignados
        if grupo.user_set.exists():
            return JsonResponse({
                'success': False,
                'error': f'El grupo {nombre} tiene usuarios asignados. Reasígnalos primero.'
            }, status=400)
        
        grupo.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Grupo {nombre} eliminado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
